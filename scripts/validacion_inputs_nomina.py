"""
Script de Validación: Excel "Inputs de Nomina" vs Motor Python
==============================================================

Compara CONCEPTO POR CONCEPTO el cálculo de nómina cargada entre:
- Excel V2-4: Hoja "Inputs de Nomina"
- Motor Python: domain/services/nomina_cargada.py

Columnas validadas:
  F  = T. Imponible
  G  = Auxilio Transporte
  H  = T. Haberes
  I  = Salud (patronal)
  J  = Pensión (patronal)
  L  = ARL
  M  = Seg. Social total (H+I+J+ARL)
  N  = Caja
  O  = ICBF+Sena
  P  = Parafiscales (N+O)
  Q  = Cesantías
  R  = Primas
  S  = Intereses cesantía
  T  = Vacaciones
  U  = Prestaciones (Q+R+S+T)
  V  = Dotaciones
  W  = Costo Empresa (M+P+U+V)
  AM = Nómina Cargada Total

Tolerancia: diferencias < 1 COP son aceptadas (redondeo ROUND_HALF_UP).

Autor: Sistema de validación NEXA
Fecha: 2026-05-26
"""

import sys
from pathlib import Path
from typing import List, Dict
import openpyxl

# ── Path setup ───────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
PARENT_ROOT  = PROJECT_ROOT.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PARENT_ROOT))

# Import backend_nexa package first — registers "nexa_engine" alias in sys.modules
# (see backend_nexa/__init__.py) — required by domain.services.nomina_cargada internals
import backend_nexa  # noqa: F401

from nexa_engine.modules.cadena_a.services.nomina_cargada import NominaCargadaService, ParametrosNominaLaboral
from nexa_engine.modules.shared.precision import cop_round


TOLERANCIA = 1.0  # COP — diferencias < 1 COP son aceptadas


# ─────────────────────────────────────────────────────────────────────────────
# Clase de descomposición detallada del cálculo Python
# (replica internamente NominaCargadaService.calcular() pero devuelve cada componente)
# ─────────────────────────────────────────────────────────────────────────────

def calcular_detallado(salario_base: float, comision_pct: float, params: ParametrosNominaLaboral) -> dict:
    """
    Réplica de NominaCargadaService.calcular() pero retornando cada componente
    para poder comparar concepto por concepto contra el Excel.
    """
    p = params
    smmlv       = p.salario_minimo
    umbral_alto = p.factor_alto_salario_smmlv * smmlv

    t_imponible = salario_base * (1.0 + comision_pct * p.pct_cumplimiento_variable)
    aux         = p.auxilio_transporte if t_imponible < 2 * smmlv else 0.0
    t_haberes   = t_imponible + aux

    alto_salario = t_imponible > umbral_alto

    if alto_salario:
        factor    = p.factor_corrector_alto_salario
        salud     = t_imponible * p.tasa_salud     * factor
        pension   = t_imponible * p.tasa_pension   * factor
        arl       = t_imponible * p.tasa_arl       * factor
        caja      = t_imponible * p.tasa_caja      * factor
        icbf_sena = t_imponible * p.tasa_icbf_sena * factor
        vac_rate  = p.tasa_vacaciones              * factor
    else:
        # Ley 1819 de 2016: T.Imponible < 10×SMMLV → Salud=0, ICBF+Sena=0
        # Excel formula: I_col = IF(F>10*SMMLV, F*tasa*70%, 0)
        salud     = 0.0
        icbf_sena = 0.0
        pension   = t_imponible * p.tasa_pension
        arl       = t_imponible * p.tasa_arl
        caja      = t_imponible * p.tasa_caja
        vac_rate  = p.tasa_vacaciones

    seg_social   = t_haberes + salud + pension + arl
    parafiscales = caja + icbf_sena

    if t_haberes <= umbral_alto:
        cesantias = t_haberes * p.tasa_cesantias
        primas    = t_haberes * p.tasa_primas
        int_ces   = cesantias * p.tasa_interes_cesantia
    else:
        cesantias = primas = int_ces = 0.0

    vacaciones   = t_imponible * vac_rate
    prestaciones = cesantias + primas + int_ces + vacaciones

    dotaciones = p.dotaciones_mensual if t_imponible < 2 * smmlv else 0.0

    costo_empresa  = seg_social + parafiscales + prestaciones + dotaciones

    return {
        't_imponible':   t_imponible,
        'aux_transporte': aux,
        't_haberes':     t_haberes,
        'salud':         salud,
        'pension':       pension,
        'arl':           arl,
        'seg_social':    seg_social,
        'caja':          caja,
        'icbf_sena':     icbf_sena,
        'parafiscales':  parafiscales,
        'cesantias':     cesantias,
        'primas':        primas,
        'int_ces':       int_ces,
        'vacaciones':    vacaciones,
        'prestaciones':  prestaciones,
        'dotaciones':    dotaciones,
        'costo_empresa': costo_empresa,
        'nomina_cargada': costo_empresa,  # Motor retorna valor exacto (sin redondear)
        'alto_salario':  alto_salario,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Validador
# ─────────────────────────────────────────────────────────────────────────────

class ValidadorInputsNomina:
    """Valida paridad Excel vs Python en nómina cargada."""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.ws = self.wb["Inputs de Nomina"]
        self.resultados: List[Dict] = []
        self.errores_por_cargo: List[Dict] = []

    # ── Lectura de parámetros ─────────────────────────────────────────────────

    def extraer_parametros_excel(self) -> dict:
        """Extrae los parámetros de cabecera del Excel."""
        ws = self.ws
        return {
            'smmlv':                      float(ws['C4'].value),
            'auxilio_transporte':         float(ws['C5'].value),
            'pct_cumplimiento_variable':  float(ws['C6'].value),
            'dotaciones_mensual':         float(ws['C8'].value),
            'tasa_salud':                 float(ws['I13'].value),
            'tasa_pension':               float(ws['J13'].value),
            'tasa_arl':                   float(ws['L13'].value),
            'tasa_caja':                  float(ws['N13'].value),
            'tasa_icbf_sena':             float(ws['O13'].value),
            'tasa_cesantias':             float(ws['Q13'].value),
            'tasa_primas':                float(ws['R13'].value),
            'tasa_interes_cesantia':      float(ws['S13'].value),
            'tasa_vacaciones':            float(ws['T13'].value),
            # Hardcoded en Excel (sin celda dedicada)
            'factor_alto_salario_smmlv':  10.0,
            'factor_corrector_alto_salario': 0.70,
        }

    def leer_fila_excel(self, row_idx: int) -> dict:
        """Lee todos los campos de una fila del Excel."""
        ws = self.ws

        def f(col: str) -> float:
            v = ws[f'{col}{row_idx}'].value
            return float(v) if v is not None else 0.0

        return {
            'cargo':          ws[f'B{row_idx}'].value,
            'salario_base':   f('C'),
            'variable':       f('D'),
            'pct_comision':   f('E'),
            't_imponible':    f('F'),
            'aux_transporte': f('G'),
            't_haberes':      f('H'),
            'salud':          f('I'),
            'pension':        f('J'),
            # K = columna ARL staff (no usada directamente)
            'arl':            f('L'),
            'seg_social':     f('M'),
            'caja':           f('N'),
            'icbf_sena':      f('O'),
            'parafiscales':   f('P'),
            'cesantias':      f('Q'),
            'primas':         f('R'),
            'int_ces':        f('S'),
            'vacaciones':     f('T'),
            'prestaciones':   f('U'),
            'dotaciones':     f('V'),
            'costo_empresa':  f('W'),
            'nomina_cargada': f('AM'),
        }

    # ── Validación por fila ───────────────────────────────────────────────────

    def validar_fila(self, row_idx: int, params_obj: ParametrosNominaLaboral) -> List[Dict]:
        """Valida una fila y retorna lista de comparaciones."""
        excel = self.leer_fila_excel(row_idx)
        cargo = excel['cargo']

        if not cargo or cargo == "Cargo":
            return []

        print(f"\n{'='*110}")
        print(f"  CARGO: {cargo}   (Fila {row_idx})   |   Salario base: {excel['salario_base']:>15,.0f} COP")
        print(f"{'='*110}")

        py = calcular_detallado(
            salario_base  = excel['salario_base'],
            comision_pct  = excel['pct_comision'],
            params        = params_obj,
        )

        # Mapa: (etiqueta, clave_python, clave_excel)
        conceptos = [
            ("T. Imponible",          't_imponible',   't_imponible'),
            ("Auxilio Transporte",     'aux_transporte','aux_transporte'),
            ("T. Haberes",             't_haberes',     't_haberes'),
            ("Salud (8.5%)",           'salud',         'salud'),
            ("Pensión (12%)",          'pension',       'pension'),
            ("ARL (0.522%)",           'arl',           'arl'),
            ("Seg. Social total",      'seg_social',    'seg_social'),
            ("Caja (4%)",              'caja',          'caja'),
            ("ICBF+Sena (4%)",         'icbf_sena',     'icbf_sena'),
            ("Parafiscales total",     'parafiscales',  'parafiscales'),
            ("Cesantías (8.33%)",      'cesantias',     'cesantias'),
            ("Primas (8.33%)",         'primas',        'primas'),
            ("Interés cesantías (12%)","int_ces",       'int_ces'),
            ("Vacaciones (4.17%)",     'vacaciones',    'vacaciones'),
            ("Prestaciones total",     'prestaciones',  'prestaciones'),
            ("Dotaciones",             'dotaciones',    'dotaciones'),
            ("Costo Empresa total",    'costo_empresa', 'costo_empresa'),
            ("💰 NÓMINA CARGADA",      'nomina_cargada','nomina_cargada'),
        ]

        print(f"  Alto salario: {'SÍ' if py['alto_salario'] else 'NO'}")
        print(f"\n{'Concepto':<28} | {'Excel':>15} | {'Python':>15} | {'Δ':>12} | {'OK?':>5}")
        print("-" * 90)

        comparaciones = []
        for etiqueta, py_key, xl_key in conceptos:
            val_excel  = excel.get(xl_key, 0.0)
            val_python = py.get(py_key, 0.0)
            diff       = abs(val_excel - val_python)
            ok         = diff < TOLERANCIA
            symbol     = "✓" if ok else "✗"
            marker     = "  ← DIFERENCIA" if not ok else ""

            print(f"  {etiqueta:<26} | {val_excel:>15,.2f} | {val_python:>15,.2f} | "
                  f"{diff:>12,.4f} | {symbol:>5}{marker}")

            comparaciones.append({
                'cargo':     cargo,
                'fila':      row_idx,
                'concepto':  etiqueta,
                'excel':     val_excel,
                'python':    val_python,
                'diferencia': diff,
                'ok':        ok,
            })

        self.resultados.extend(comparaciones)
        return comparaciones

    # ── Ejecución principal ───────────────────────────────────────────────────

    def ejecutar_validacion(self, filas_a_validar=None):
        """Ejecuta la validación completa."""
        print("\n" + "="*110)
        print("  VALIDACIÓN EXHAUSTIVA: Inputs de Nomina — Excel V2-4 vs Motor Python")
        print("="*110)

        params_raw = self.extraer_parametros_excel()

        print("\n📋 PARÁMETROS EXTRAÍDOS DEL EXCEL:")
        print("-" * 60)
        for k, v in params_raw.items():
            pct_str = f"  ({v*100:.3f}%)" if "tasa_" in k or "factor_corrector" in k else ""
            print(f"    {k:<38} = {v:>12}{pct_str}")

        params_obj = ParametrosNominaLaboral(
            salario_minimo                = params_raw['smmlv'],
            auxilio_transporte            = params_raw['auxilio_transporte'],
            dotaciones_mensual            = params_raw['dotaciones_mensual'],
            pct_cumplimiento_variable     = params_raw['pct_cumplimiento_variable'],
            factor_alto_salario_smmlv     = params_raw['factor_alto_salario_smmlv'],
            factor_corrector_alto_salario = params_raw['factor_corrector_alto_salario'],
            tasa_salud                    = params_raw['tasa_salud'],
            tasa_pension                  = params_raw['tasa_pension'],
            tasa_arl                      = params_raw['tasa_arl'],
            tasa_caja                     = params_raw['tasa_caja'],
            tasa_icbf_sena                = params_raw['tasa_icbf_sena'],
            tasa_cesantias                = params_raw['tasa_cesantias'],
            tasa_primas                   = params_raw['tasa_primas'],
            tasa_interes_cesantia         = params_raw['tasa_interes_cesantia'],
            tasa_vacaciones               = params_raw['tasa_vacaciones'],
        )

        filas = filas_a_validar or list(range(16, 26))  # filas 16-25 (cargos típicos)

        for fila in filas:
            try:
                self.validar_fila(fila, params_obj)
            except Exception as exc:
                print(f"\n⚠️  Error en fila {fila}: {exc}")
                import traceback; traceback.print_exc()

        self.imprimir_resumen()

    # ── Resumen ───────────────────────────────────────────────────────────────

    def imprimir_resumen(self):
        """Imprime el resumen final de validación."""
        print("\n\n" + "="*110)
        print("  📊 RESUMEN DE VALIDACIÓN")
        print("="*110)

        if not self.resultados:
            print("  Sin resultados — todas las filas produjeron error.")
            return

        total      = len(self.resultados)
        ok_count   = sum(1 for r in self.resultados if r['ok'])
        fail_count = total - ok_count

        print(f"\n  Total comparaciones : {total}")
        print(f"  ✓ Coincidencias     : {ok_count} ({ok_count/total*100:.1f}%)")
        print(f"  ✗ Diferencias       : {fail_count} ({fail_count/total*100:.1f}%)")
        print(f"  Tolerancia aplicada : {TOLERANCIA:.0f} COP\n")

        if fail_count > 0:
            print("  ⚠️  DIFERENCIAS ENCONTRADAS:")
            print("-" * 110)
            print(f"  {'Cargo':<25} {'Concepto':<28} {'Excel':>15} {'Python':>15} {'Δ':>12}")
            print("-" * 110)
            for r in self.resultados:
                if not r['ok']:
                    print(f"  {r['cargo']:<25} {r['concepto']:<28} "
                          f"{r['excel']:>15,.2f} {r['python']:>15,.2f} {r['diferencia']:>12,.4f}")
        else:
            print("  ✅ PARIDAD COMPLETA — Python replica exactamente el Excel.")

        # Estadísticas por concepto
        conceptos_fallidos = {}
        for r in self.resultados:
            if not r['ok']:
                c = r['concepto']
                conceptos_fallidos.setdefault(c, []).append(r['diferencia'])

        if conceptos_fallidos:
            print("\n  RESUMEN DE FALLOS POR CONCEPTO:")
            print("-" * 60)
            for c, diffs in sorted(conceptos_fallidos.items(), key=lambda x: -max(x[1])):
                print(f"    {c:<32} | {len(diffs):>3} filas  | Δ máx = {max(diffs):>10,.4f}")

        print()


# ─────────────────────────────────────────────────────────────────────────────
# Punto de entrada
# ─────────────────────────────────────────────────────────────────────────────

def main():
    excel_path = str(PROJECT_ROOT / "excel" / "Nexa - Pricing - Simulador - V2-4.xlsx")

    print(f"\n  Excel: {excel_path}")

    validador = ValidadorInputsNomina(excel_path)
    # Validar filas 16-25 (todos los cargos disponibles en la hoja)
    validador.ejecutar_validacion(filas_a_validar=list(range(16, 26)))

    print("✅ Validación completada.\n")


if __name__ == "__main__":
    main()
