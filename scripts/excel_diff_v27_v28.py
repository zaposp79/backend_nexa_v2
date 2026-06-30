"""
scripts/excel_diff_v27_v28.py
=============================
Core Stage-1 deliverable: diff celda-por-celda entre Excel V2-7 y V2-8.

El backend está certificado en paridad V2-7, por lo que el gap de paridad V2-8
es EXACTAMENTE lo que cambió entre ambos workbooks. Este script lo aísla.

Gates obligatorios (ajustes aprobados):
  1. Verifica SHA256 del V2-7 actual contra el certificado. Mismatch ->
     KILL_SWITCH V27_DRIFT_DETECTED.
  2. Reporta tres conjuntos de hojas: SHEETS_BOTH, SHEETS_ONLY_V28,
     SHEETS_ONLY_V27.
  3. Clasifica cada celda cambiada en:
        SEMANTIC_CHANGE   -> fórmula cambió, o valor cacheado cambió > tol
        FORMAT_ONLY       -> mismo valor/fórmula; difiere solo formato
        COMMENT_OR_LABEL  -> cambio en texto/label/comentario (string no numérico)
        STRUCTURAL        -> celda agregada/eliminada
     Solo SEMANTIC_CHANGE (y FORMAT_ONLY si afecta presentación de vista) se
     proponen como candidatos a cambio en Stage 2.

Salidas:
  - docs/refactor/excel_v27_v28_diff.json
  - docs/refactor/excel_v27_v28_diff.md

Uso:
    PYTHONPATH=$(pwd) python scripts/excel_diff_v27_v28.py
"""
from __future__ import annotations

import json
import re
import sys
import warnings
from collections import Counter
from datetime import datetime, timezone

import openpyxl
from openpyxl.utils import column_index_from_string, coordinate_to_tuple

from scripts.excel_map_common import (
    BACKEND_ROOT,
    EXCEL_V27_PATH,
    EXCEL_V28_PATH,
    EXPECTED_V27_SHA256,
    SHEET_TYPE,
    TOL_MONETARY,
    sha256_of,
)

warnings.filterwarnings("ignore")

OUT_JSON = BACKEND_ROOT / "docs" / "refactor" / "excel_v27_v28_diff.json"
OUT_MD = BACKEND_ROOT / "docs" / "refactor" / "excel_v27_v28_diff.md"

# Tolerancia numérica genérica para "valor cacheado cambió".
# Conservadora: cualquier delta absoluto > esto en un número se marca SEMANTIC.
_NUM_TOL = TOL_MONETARY


def _check_v27_drift() -> str:
    actual = sha256_of(EXCEL_V27_PATH)
    if actual != EXPECTED_V27_SHA256:
        raise SystemExit(
            f"KILL_SWITCH: V27_DRIFT_DETECTED — V2-7 actual sha256={actual} "
            f"!= certificado={EXPECTED_V27_SHA256}. El backend está certificado "
            f"contra el workbook certificado; abortar diff."
        )
    return actual


def _is_number(v: object) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _is_formula(v: object) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _norm_formula(v: object) -> str:
    return "".join(str(v).split())


# Matchea referencias de celda A1 (con $ opcionales), NO precedidas por
# una letra (para no romper nombres de función como LOG10 o ABS).
_CELL_REF_RE = re.compile(r"(?<![A-Za-z0-9_])(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _to_r1c1(formula: str, anchor_coord: str) -> str:
    """Convierte una fórmula A1 a R1C1 relativa a su celda ancla.

    Hace que las referencias relativas sean invariantes ante inserción/
    borrado uniforme de filas/columnas (colapsa ruido de layout shift).
    Referencias absolutas ($) se mantienen como índice absoluto.
    Conversión best-effort: ignora prefijos de hoja y literales de string.
    """
    arow, acol = coordinate_to_tuple(anchor_coord)

    def repl(m: re.Match) -> str:
        col_abs, col_letters, row_abs, row_digits = m.groups()
        col = column_index_from_string(col_letters)
        row = int(row_digits)
        c_part = f"C{col}" if col_abs else f"C[{col - acol}]"
        r_part = f"R{row}" if row_abs else f"R[{row - arow}]"
        return r_part + c_part

    return _CELL_REF_RE.sub(repl, formula)


def _formula_logic_equal(f7: str, f8: str, coord: str) -> bool:
    """Igualdad de lógica exacta: idéntica en A1 normalizado o en R1C1."""
    if _norm_formula(f7) == _norm_formula(f8):
        return True
    try:
        r7 = _to_r1c1(_norm_formula(f7), coord)
        r8 = _to_r1c1(_norm_formula(f8), coord)
        return r7 == r8
    except Exception:
        return False


# Masca prefijos de hoja ('Hoja'! o Hoja!), rangos full-column (A:A, $D:$D),
# rangos de celda (A1:B2) y celdas sueltas (A1) -> '@'.
_SHEET_PREFIX_RE = re.compile(r"('[^']*'|[A-Za-z_][A-Za-z0-9_. ]*)!")
_FULLCOL_RE = re.compile(r"\$?[A-Z]{1,3}:\$?[A-Z]{1,3}")
_RANGE_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+")
_CELL_RE = re.compile(r"(?<![A-Za-z0-9_])\$?[A-Z]{1,3}\$?\d+")


def _formula_skeleton(formula: str) -> str:
    """Estructura de la fórmula con TODAS las referencias enmascaradas a '@'.

    Si dos fórmulas tienen el mismo skeleton, difieren solo en a qué celdas
    apuntan (reference retarget por shift de layout), no en lógica.
    """
    s = "".join(str(formula).split())
    s = _SHEET_PREFIX_RE.sub("@!", s)
    s = _FULLCOL_RE.sub("@", s)
    s = _RANGE_RE.sub("@", s)
    s = _CELL_RE.sub("@", s)
    return s


# Conjuntos de clasificación.
#   Parity-relevant (lógica): FORMULA_CHANGED, FORMULA_ADDED, FORMULA_REMOVED,
#                             CONSTANT_CHANGED
#   Ruido (input-driven):     VALUE_ONLY  (misma fórmula, distinto cacheado)
#   Texto:                    LABEL_CHANGED
#   Datos:                    DATA_STRUCTURAL (celda literal agregada/eliminada)
# Candidatos REALES a paridad (lógica accionable):
PARITY_KINDS = frozenset({"FORMULA_LOGIC_CHANGED", "CONSTANT_CHANGED"})
# Estructural de fórmula: contaminado por shifts de layout, validar vía runner:
FORMULA_STRUCTURAL_KINDS = frozenset(
    {"REFERENCE_RETARGET", "FORMULA_ADDED", "FORMULA_REMOVED"}
)


def _classify(
    f7: object, f8: object, v7: object, v8: object, coord: str,
) -> str | None:
    """Clasifica el cambio de una celda. None si no hubo cambio relevante.

    Clave de paridad: el LOGIC delta vive en la fórmula, no en el valor
    cacheado (que depende del deal de ejemplo cargado en cada workbook).
    Las fórmulas se comparan en R1C1 para ignorar shifts de layout.
    """
    f7_is = _is_formula(f7)
    f8_is = _is_formula(f8)

    # --- Caso fórmula en al menos un lado ---
    if f7_is or f8_is:
        if f7_is and f8_is:
            if _formula_logic_equal(f7, f8, coord):
                # Misma lógica exacta: diff de valor es input-driven -> ruido
                return None
            # Difieren textualmente: ¿solo retarget o lógica real?
            if _formula_skeleton(f7) == _formula_skeleton(f8):
                return "REFERENCE_RETARGET"
            return "FORMULA_LOGIC_CHANGED"
        # Una celda tiene fórmula y la otra no
        return "FORMULA_ADDED" if f8_is else "FORMULA_REMOVED"

    # --- Ambos lados son literales (no fórmulas) ---
    if _is_number(v7) and _is_number(v8):
        if abs(float(v7) - float(v8)) > _NUM_TOL:
            return "CONSTANT_CHANGED"
        return None
    # texto / mixto
    a = v7 if v7 is not None else None
    b = v8 if v8 is not None else None
    if a != b:
        return "LABEL_CHANGED"
    return None


def _scan_sheet(ws_f7, ws_v7, ws_f8, ws_v8) -> list[dict]:
    """Diff de una hoja presente en ambos workbooks."""
    max_row = max(ws_f7.max_row, ws_f8.max_row)
    max_col = max(ws_f7.max_column, ws_f8.max_column)
    rows: list[dict] = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell7f = ws_f7.cell(row=r, column=c)
            cell8f = ws_f8.cell(row=r, column=c)
            f7, f8 = cell7f.value, cell8f.value
            v7 = ws_v7.cell(row=r, column=c).value
            v8 = ws_v8.cell(row=r, column=c).value

            present7 = f7 is not None or v7 is not None
            present8 = f8 is not None or v8 is not None
            if not present7 and not present8:
                continue

            kind = _classify(f7, f8, v7, v8, cell8f.coordinate)
            if kind is None:
                continue
            # Literal (no fórmula) agregado/eliminado = dato, no lógica.
            if kind == "LABEL_CHANGED" and present7 != present8:
                if not (_is_formula(f7) or _is_formula(f8)):
                    kind = "DATA_STRUCTURAL"
            rows.append({
                "coord": cell8f.coordinate,
                "kind": kind,
                "f_v27": f7, "f_v28": f8, "v_v27": v7, "v_v28": v8,
            })
    return rows


def _logic_transitions(diffs: list[dict]) -> list[dict]:
    """Agrupa FORMULA_LOGIC_CHANGED por (skeleton_v27 -> skeleton_v28).

    Devuelve transiciones distintas con conteo y celda de ejemplo. Esta es la
    lista accionable real para Stage 2: N transiciones, no miles de celdas.
    """
    groups: dict[tuple[str, str], dict] = {}
    for c in diffs:
        if c["kind"] != "FORMULA_LOGIC_CHANGED":
            continue
        key = (_formula_skeleton(c["f_v27"]), _formula_skeleton(c["f_v28"]))
        g = groups.get(key)
        if g is None:
            groups[key] = {
                "skeleton_v27": key[0],
                "skeleton_v28": key[1],
                "count": 1,
                "example_coord": c["coord"],
                "example_v27": c["f_v27"],
                "example_v28": c["f_v28"],
            }
        else:
            g["count"] += 1
    return sorted(groups.values(), key=lambda g: g["count"], reverse=True)


def main() -> int:
    include_full = "--full" in sys.argv  # incluir dump per-celda (genera JSON enorme)
    v27_hash = _check_v27_drift()

    wb_f7 = openpyxl.load_workbook(EXCEL_V27_PATH, data_only=False)
    wb_v7 = openpyxl.load_workbook(EXCEL_V27_PATH, data_only=True)
    wb_f8 = openpyxl.load_workbook(EXCEL_V28_PATH, data_only=False)
    wb_v8 = openpyxl.load_workbook(EXCEL_V28_PATH, data_only=True)

    sheets7 = set(wb_f7.sheetnames)
    sheets8 = set(wb_f8.sheetnames)
    sheets_both = sorted(sheets7 & sheets8)
    sheets_only_v28 = sorted(sheets8 - sheets7)
    sheets_only_v27 = sorted(sheets7 - sheets8)

    per_sheet: dict[str, dict] = {}
    total_transitions = 0
    for name in sheets_both:
        diffs = _scan_sheet(
            wb_f7[name], wb_v7[name], wb_f8[name], wb_v8[name],
        )
        counts = Counter(d["kind"] for d in diffs)
        transitions = _logic_transitions(diffs)
        total_transitions += len(transitions)
        per_sheet[name] = {
            "sheet_type": SHEET_TYPE.get(name, "UNKNOWN"),
            "total_changes": len(diffs),
            "by_kind": dict(counts),
            "logic_transitions": transitions,
            "changes": diffs,
        }

    result = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "v27_path": str(EXCEL_V27_PATH.relative_to(BACKEND_ROOT)),
        "v28_path": str(EXCEL_V28_PATH.relative_to(BACKEND_ROOT)),
        "v27_sha256": v27_hash,
        "v27_drift_gate": "PASS",
        "sheets_both": sheets_both,
        "sheets_only_v28": sheets_only_v28,
        "sheets_only_v27": sheets_only_v27,
        "total_logic_transitions": total_transitions,
        "per_sheet": per_sheet,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_MD.write_text(_render_md(result))

    # JSON compacto por defecto: conserva by_kind + logic_transitions +
    # constant_changes; descarta el dump per-celda (retarget/add/rem) que
    # infla el archivo a ~31MB. --full lo incluye todo.
    json_result = {**result, "per_sheet": {}}
    for name, ps in result["per_sheet"].items():
        compact = {k: v for k, v in ps.items() if k != "changes"}
        compact["constant_changes"] = [
            c for c in ps["changes"] if c["kind"] == "CONSTANT_CHANGED"
        ]
        if include_full:
            compact["changes"] = ps["changes"]
        json_result["per_sheet"][name] = compact
    OUT_JSON.write_text(json.dumps(json_result, ensure_ascii=False, indent=2, default=str))

    # CLI summary
    print(f"V2-7 drift gate: PASS (sha256={v27_hash[:12]}…)")
    print(f"SHEETS_BOTH: {len(sheets_both)} | ONLY_V28: {sheets_only_v28} | ONLY_V27: {sheets_only_v27}")
    print("\nLOGIC=fórmula con skeleton distinto (REAL) | CONST=constante | "
          "RETARGET/ADD/REM=shift de layout (ruido):")
    grand = Counter()
    for name in sheets_both:
        bk = per_sheet[name]["by_kind"]
        grand.update(bk)
        parity = sum(bk.get(k, 0) for k in PARITY_KINDS)
        struct = sum(bk.get(k, 0) for k in FORMULA_STRUCTURAL_KINDS)
        if parity or struct:
            flag = " <<<" if parity else ""
            print(f"  [{per_sheet[name]['sheet_type']:5}] {name:32} "
                  f"LOGIC={bk.get('FORMULA_LOGIC_CHANGED', 0):4} "
                  f"CONST={bk.get('CONSTANT_CHANGED', 0):4} | PARITY={parity:4} "
                  f"(ruido shift={struct}){flag}")
    parity_total = sum(grand.get(k, 0) for k in PARITY_KINDS)
    print(f"\nTOTAL por tipo: {dict(grand)}")
    print(f"\n>>> DISTINCT LOGIC TRANSITIONS (changeset accionable Stage 2): "
          f"{total_transitions}")
    print(f">>> PARITY CANDIDATES (celdas): {parity_total} "
          f"(LOGIC={grand.get('FORMULA_LOGIC_CHANGED', 0)}, "
          f"CONST={grand.get('CONSTANT_CHANGED', 0)})")
    print(f"    Ruido de shift: RETARGET={grand.get('REFERENCE_RETARGET', 0)}, "
          f"F_ADD={grand.get('FORMULA_ADDED', 0)}, F_REM={grand.get('FORMULA_REMOVED', 0)}, "
          f"DATA={grand.get('DATA_STRUCTURAL', 0)}, LABEL={grand.get('LABEL_CHANGED', 0)}")
    print(f"\n→ {OUT_JSON.relative_to(BACKEND_ROOT)}")
    print(f"→ {OUT_MD.relative_to(BACKEND_ROOT)}")
    return 0


def _render_md(result: dict) -> str:
    lines = [
        "# Excel V2-7 → V2-8 Diff (Stage 1)",
        f"**Generado:** {result['generated_at']}",
        f"**V2-7:** `{result['v27_path']}` (sha256 `{result['v27_sha256'][:16]}…`, drift gate {result['v27_drift_gate']})",
        f"**V2-8:** `{result['v28_path']}`",
        "",
        "## Conjuntos de hojas",
        f"- **SHEETS_BOTH** ({len(result['sheets_both'])}): {', '.join(result['sheets_both'])}",
        f"- **SHEETS_ONLY_V28** ({len(result['sheets_only_v28'])}): {', '.join(result['sheets_only_v28']) or '—'} "
        f"→ candidatos **MISSING_IN_BACKEND**",
        f"- **SHEETS_ONLY_V27** ({len(result['sheets_only_v27'])}): {', '.join(result['sheets_only_v27']) or '—'} "
        f"→ **requiere decisión humana**",
        "",
        "## Metodología de clasificación",
        "",
        "Ambos workbooks llevan un deal de ejemplo distinto cargado, por lo que",
        "comparar valores cacheados es ruido. El delta de **lógica** vive en la",
        "fórmula. Además, V2-8 insertó/eliminó filas en varias hojas, lo que",
        "**retarget**ea referencias absolutas (`'Hoja'!$A$51`→`$A$53`) sin",
        "cambiar la lógica. Para aislar la señal real se usa el **skeleton** de",
        "la fórmula (toda referencia enmascarada a `@`):",
        "",
        "- **PARITY CANDIDATES** (lógica real, accionables en Stage 2):",
        "  - `FORMULA_LOGIC_CHANGED` — skeleton distinto (cambió funciones/",
        "    operadores/estructura, no solo a qué celda apunta).",
        "  - `CONSTANT_CHANGED` — literal numérico hardcodeado cambió.",
        "- **Ruido de shift de layout** (NO candidatos; validar vía parity",
        "  runner contra valores cacheados V2-8):",
        "  - `REFERENCE_RETARGET` — misma skeleton, distinta celda destino.",
        "  - `FORMULA_ADDED` / `FORMULA_REMOVED` — fórmula aparece/desaparece",
        "    en una coordenada por desplazamiento de bloque.",
        "  - `DATA_STRUCTURAL` / `LABEL_CHANGED` / `VALUE_ONLY` (excluido).",
        "",
        "> **Limitación conocida:** un diff por coordenada no puede separar con",
        "> 100% de certeza un FORMULA_ADDED/REMOVED real de uno por shift sin",
        "> alinear filas hoja-por-hoja. Por eso las hojas-motor de grilla grande",
        "> (`Nomina Loaded`, `Costo Cadena C`, `Costo Fijo`, `Pólizas - Costo",
        "> Financiacion`) se validan numéricamente vía runner, no por este diff.",
        "",
        "## Resumen de PARITY CANDIDATES por hoja (lógica real)",
        "",
        "| Hoja | Tipo | FORMULA_LOGIC_CHANGED | CONSTANT_CHANGED | Parity total | Ruido shift |",
        "|------|------|----------------------:|-----------------:|-------------:|------------:|",
    ]
    for name in result["sheets_both"]:
        ps = result["per_sheet"][name]
        bk = ps["by_kind"]
        parity = sum(bk.get(k, 0) for k in PARITY_KINDS)
        struct = sum(bk.get(k, 0) for k in FORMULA_STRUCTURAL_KINDS)
        if not parity and not struct:
            continue
        lines.append(
            f"| {name} | {ps['sheet_type']} | "
            f"{bk.get('FORMULA_LOGIC_CHANGED', 0)} | "
            f"{bk.get('CONSTANT_CHANGED', 0)} | {parity} | {struct} |"
        )
    lines += [
        "",
        f"## Changeset accionable: {result['total_logic_transitions']} transiciones de fórmula distintas",
        "",
        "Cada transición = un patrón de fórmula que cambió (skeleton V2-7 →",
        "skeleton V2-8), con cuántas celdas la repiten y una celda de ejemplo.",
        "Esta es la lista de trabajo real para Stage 2.",
        "",
    ]
    for name in result["sheets_both"]:
        ps = result["per_sheet"][name]
        trans = ps.get("logic_transitions", [])
        if not trans:
            continue
        lines.append(f"### {name} ({ps['sheet_type']}) — {len(trans)} transiciones distintas")
        lines.append("")
        for t in trans:
            lines.append(
                f"- **{t['example_coord']}** ×{t['count']} celdas"
            )
            lines.append(f"  - V2-7: `{str(t['example_v27'])[:160]}`")
            lines.append(f"  - V2-8: `{str(t['example_v28'])[:160]}`")
        lines.append("")
    # Constantes cambiadas (resumen por hoja)
    lines += ["", "## CONSTANT_CHANGED por hoja (literales numéricos)", ""]
    for name in result["sheets_both"]:
        ps = result["per_sheet"][name]
        consts = [c for c in ps["changes"] if c["kind"] == "CONSTANT_CHANGED"]
        if not consts:
            continue
        lines.append(f"### {name} ({ps['sheet_type']}) — {len(consts)} constantes")
        lines.append("")
        lines.append("| Celda | V2-7 | V2-8 |")
        lines.append("|-------|------|------|")
        for c in consts[:30]:
            lines.append(f"| `{c['coord']}` | {c['v_v27']} | {c['v_v28']} |")
        if len(consts) > 30:
            lines.append(f"| … | _{len(consts) - 30} más_ | … |")
        lines.append("")
    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    sys.exit(main())
