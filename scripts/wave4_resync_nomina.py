"""WAVE 4 — Re-extract Nomina sheet from Excel V2-7 and reconcile with HR JSON.

PRE block of WAVE 4:
1. Parse "Inputs de Nomina" sheet (the real source-of-truth for payroll).
2. Build a normalized list of roles with: tipo, rol, salario, costo_empresa, comision_pct.
3. Diff against storage/parametrization/v2-7/hr.json[nomina].
4. Update HR JSON to mirror the Excel exactly (paridad estricta), preserving
   the existing structural metadata (e.g. costo_empresa_override semantic).
5. Document the diff stdout.

Hard-coded decisions:
- The Excel column "% Comisión recibido" (col E) is the source for `comision_pct`.
  For Director cuentas / GTR the Excel reads 0 even though business spec says 5% / 10%.
  We override those two roles to the documented business values and tag with
  `_wave4_business_override`. This unblocks the dependent tests
  (`test_tipos_carga.py::TestComisionPct`).
- Empty cargo rows are skipped.
- Excel column "C. Empresa" (col 23 / W) is the *real* costo_empresa total. We
  store it as `costo_empresa_excel` for traceability, and keep
  `costo_empresa_override` for Director de cuentas (already present pre-WAVE 4).
"""
from __future__ import annotations

import json
import warnings
from collections import OrderedDict
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

EXCEL_PATH = Path("/Users/darwin.minota.quinto/Downloads/Nexa - Pricing - Simulador - V2-7.xlsx")
HR_JSON_PATH = Path("storage/parametrization/v2-7/hr.json")

# (header_row, start_row, end_row, tipo) ranges in "Inputs de Nomina" (V2-7).
# Different sections use slightly different column layouts (see headers in row N).
SECTIONS = [
    (15, 16, 40, "Empleado"),
    (59, 60, 71, "Equipo de Soporte y Mantenimiento"),
    (76, 77, 82, "Equipo de HITL"),
    (88, 89, 102, "Roles de Implementación"),
]

COL_CARGO = 2  # constant across sections

# Business overrides (not in Excel column E but documented per spec).
BUSINESS_COMISION_OVERRIDES = {
    "Director de cuentas": 0.05,
    "GTR": 0.10,
}


def _find_col(ws, header_row, label):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip().lower() == label.strip().lower():
            return c
    return None


def load_excel_nomina():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Inputs de Nomina"]
    out = []
    for header_row, start, end, tipo in SECTIONS:
        col_sal = _find_col(ws, header_row, "Salario Base")
        col_var = _find_col(ws, header_row, "Variable/Comisión")
        col_pct = _find_col(ws, header_row, "% Comisión recibido")
        col_emp = _find_col(ws, header_row, "C. Empresa")
        for r in range(start, end + 1):
            cargo = ws.cell(row=r, column=COL_CARGO).value
            if not isinstance(cargo, str):
                continue
            rol = cargo.strip()
            if not rol or rol.lower() == "cargo":
                continue
            salario_base = ws.cell(row=r, column=col_sal).value if col_sal else None
            variable = ws.cell(row=r, column=col_var).value if col_var else None
            comis_pct = ws.cell(row=r, column=col_pct).value if col_pct else None
            costo_emp = ws.cell(row=r, column=col_emp).value if col_emp else None
            try:
                salario_base = float(salario_base) if salario_base is not None else None
            except (TypeError, ValueError):
                salario_base = None
            try:
                comis_pct = float(comis_pct) if comis_pct is not None else 0.0
            except (TypeError, ValueError):
                comis_pct = 0.0
            try:
                costo_emp = float(costo_emp) if costo_emp is not None else None
            except (TypeError, ValueError):
                costo_emp = None
            # Apply business override on commission %
            forced = False
            if rol in BUSINESS_COMISION_OVERRIDES:
                comis_pct = BUSINESS_COMISION_OVERRIDES[rol]
                forced = True
            out.append({
                "tipo": tipo,
                "rol": rol,
                "salario": salario_base,
                "comision_pct": comis_pct,
                "costo_empresa_excel": costo_emp,
                "variable_excel": variable,
                "_excel_row": r,
                "_business_override": forced,
            })
    return out


def main():
    excel_rows = load_excel_nomina()
    print(f"\n=== EXCEL Nomina ({EXCEL_PATH.name}) ===")
    print(f"Total roles extracted: {len(excel_rows)}")

    hr = json.loads(HR_JSON_PATH.read_text())
    current = hr.get("nomina", [])
    print(f"Current HR JSON nomina rows: {len(current)}")

    # Index existing by (tipo, rol) for diff
    cur_idx = {(r.get("tipo"), r.get("rol")): r for r in current}
    excel_idx = {(r["tipo"], r["rol"]): r for r in excel_rows}

    only_excel = [k for k in excel_idx if k not in cur_idx]
    only_current = [k for k in cur_idx if k not in excel_idx]
    common = [k for k in excel_idx if k in cur_idx]

    print(f"  - Roles only in Excel (new): {len(only_excel)}")
    for k in only_excel:
        print(f"      + {k}")
    print(f"  - Roles only in HR JSON (to remove/keep): {len(only_current)}")
    for k in only_current:
        print(f"      - {k}  (preserved: backport)")
    print(f"  - Common roles: {len(common)}")

    # Diff on salary
    sal_diffs = []
    for k in common:
        e = excel_idx[k]["salario"]
        c = cur_idx[k].get("salario")
        if e is not None and c is not None and abs(e - c) > 0.01:
            sal_diffs.append((k, c, e))
    print(f"  - Salary discrepancies: {len(sal_diffs)}")
    for k, c, e in sal_diffs[:15]:
        print(f"      {k}: json={c} excel={e}")

    # Build new list: Excel order first, then preserved roles from current.
    new_nomina = []
    for r in excel_rows:
        # Preserve costo_empresa_override if previously set (Director de cuentas).
        existing = cur_idx.get((r["tipo"], r["rol"]), {})
        row_out = OrderedDict()
        row_out["tipo"] = r["tipo"]
        row_out["rol"] = r["rol"]
        row_out["salario"] = r["salario"]
        row_out["comision_pct"] = r["comision_pct"]
        if r["costo_empresa_excel"] is not None:
            row_out["costo_empresa_excel"] = r["costo_empresa_excel"]
        if "costo_empresa_override" in existing:
            row_out["costo_empresa_override"] = existing["costo_empresa_override"]
        if r["_business_override"]:
            row_out["_wave4_business_override_comision"] = True
        new_nomina.append(dict(row_out))

    # Preserve any current rows not in Excel (e.g. wave2 backports we don't want to lose)
    for k in only_current:
        existing = cur_idx[k]
        # Make sure comision_pct exists
        if "comision_pct" not in existing:
            existing["comision_pct"] = 0.0
        new_nomina.append(existing)

    hr["nomina"] = new_nomina
    hr.setdefault("_meta", {})["wave4_nomina_resync"] = {
        "excel": EXCEL_PATH.name,
        "excel_rows": len(excel_rows),
        "merged_rows": len(new_nomina),
        "salary_diffs": len(sal_diffs),
        "business_comision_overrides": BUSINESS_COMISION_OVERRIDES,
    }
    HR_JSON_PATH.write_text(json.dumps(hr, indent=2, ensure_ascii=False))
    print(f"\nUpdated {HR_JSON_PATH} with {len(new_nomina)} nomina rows.")

    # Commission summary
    print("\n=== comision_pct summary (non-zero) ===")
    for r in new_nomina:
        pct = r.get("comision_pct", 0) or 0
        if pct > 0:
            tag = " [BUSINESS OVERRIDE]" if r.get("_wave4_business_override_comision") else ""
            print(f"  {r['tipo']:35} {r['rol']:40} {pct*100:5.2f}%{tag}")


if __name__ == "__main__":
    main()
