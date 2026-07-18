"""
scripts/v28_input_mapper.py — Stage 1.5 Paso A  (CHECKPOINT A, read-only)
==========================================================================
Lee celdas de input del deal en V2-8, las mapea a request.json, clasifica
cada delta y produce docs/refactor/v28_input_mapping.md.

NO modifica request.json. Detener en CHECKPOINT A.

Clasificaciones:
  MATCH                — valores idénticos (o equivalentes en tolerancia)
  VALUE_UPDATE         — valor V2-8 difiere, req.json key existe → candidato Paso B
  SCALE_MISMATCH       — difieren solo por factor potencia de 10 → revisar unidad
  STRUCTURE_EXTENSION  — V2-8 tiene input, no existe key en req.json, PERO
                         puede agregarse dentro de un objeto top-level existente
                         (sin nueva key top-level, sin cambio de tipo)
  REQUEST_STRUCTURE_GAP — diferencia estructural profunda; no parcheable como
                         VALUE_UPDATE solo
  EXCEL_LIKELY_BUG     — fuente Panel y fuente Tasas TRM difieren en V2-8 mismo;
                         no copiar ciegamente
  UNKNOWN_SOURCE       — valor existe en req.json pero no se puede rastrear a
                         ninguna celda V2-8

Uso:
    cd /Users/darwin.minota.quinto/Projects/NEXA
    PYTHONPATH=. python backend_nexa/scripts/v28_input_mapper.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

from scripts.excel_map_common import (
    BACKEND_ROOT,
    EXCEL_V28_PATH,
    REQUEST_PATH,
    is_scale_mismatch,
)

OUT_MD = BACKEND_ROOT / "docs" / "refactor" / "v28_input_mapping.md"

# ── Classification labels ─────────────────────────────────────────────────────
MATCH = "MATCH"
VALUE_UPDATE = "VALUE_UPDATE"
SCALE = "SCALE_MISMATCH"
STRUCT_EXT = "STRUCTURE_EXTENSION"
STRUCT_GAP = "REQUEST_STRUCTURE_GAP"
EXCEL_BUG = "EXCEL_LIKELY_BUG"
UNKNOWN_SRC = "UNKNOWN_SOURCE"

TOL_NUM: float = 1e-9

# ── Panel de Control General → request.json path map ─────────────────────────
# req=None + tag=STRUCT_EXT  → input exists in V2-8 but key missing in req.json,
#                              can be ADDED to an existing top-level object.
PANEL_MAP: list[dict] = [
    {"cell": "C5",  "label": "Servicio",                 "req": "datos_operativos.servicio"},
    {"cell": "C6",  "label": "Cliente",                  "req": "datos_operativos.cliente"},
    {
        "cell": "C7", "label": "Antigüedad", "req": None,
        "tag": STRUCT_EXT,
        "notes": "Nueva key datos_operativos.antiguedad; no existe hoy",
    },
    {"cell": "C8",  "label": "Tipo cliente",             "req": "datos_operativos.tipo_cliente"},
    {
        "cell": "C9", "label": "Período pago (días)", "req": None,
        "tag": STRUCT_EXT,
        "notes": "Nueva key datos_operativos.periodo_pago; no existe hoy",
    },
    {"cell": "C10", "label": "Fecha Inicio",             "req": "datos_operativos.fecha_inicio",
     "notes": "IMPORTANTE: afecta gate SUMIFS P&G en V2-8 (ver Nota Fechas)"},
    {"cell": "C11", "label": "Duración meses",           "req": "datos_operativos.duracion_meses"},
    {"cell": "C12", "label": "Ciudad",                   "req": "datos_operativos.ciudad"},
    {"cell": "C13", "label": "Sede",                     "req": "datos_operativos.sede",
     "suffix": True,
     "notes": "V2-8='Bogota - Toberin'; req='Toberin' (sufijo); motor usa req.sede"},
    {"cell": "C16", "label": "Tarifa diaria cap.",       "req": "datos_operativos.tarifa_diaria_capacitacion"},
    {"cell": "C17", "label": "Crucero",                  "req": "datos_operativos.crucero"},
    {"cell": "C18", "label": "Horas formación mes",      "req": "datos_operativos.horas_formacion_mes"},
    {"cell": "C19", "label": "% Ausentismo",             "req": "datos_operativos.pct_ausentismo"},
    {"cell": "C20", "label": "% Rotación",               "req": "datos_operativos.pct_rotacion"},
    {"cell": "C21", "label": "Considera financiación",   "req": "datos_operativos.cons_costo_de_financiacion",
     "bool": True, "notes": "V2-8='No'→False; req=True"},
    {"cell": "C28", "label": "Ciudad proporción (Bgta)", "req": "datos_operativos.ciudades_recurso[0].proporcion"},
    {
        "cell": "C34", "label": "ICA", "req": "datos_operativos.tasa_ica",
        "notes": ("cross-check Tasas: Bogota base=Tasas!B37=0.00966; "
                  "Bogota total(+bomberos)=Tasas!F37=0.01966; "
                  "Panel=0.01; req=0.0097 — NINGUNO coincide exactamente"),
    },
    {"cell": "C35", "label": "GMF",                      "req": "datos_operativos.tasa_gmf"},
    {"cell": "C63", "label": "Margen obj Cadena A",      "req": "reglas_negocio.margen_objetivo_cadena_a"},
    {
        "cell": "D63", "label": "Margen obj Cadena B", "req": None,
        "tag": STRUCT_EXT,
        "notes": "Nueva key reglas_negocio.margen_objetivo_cadena_b=0.30; req tiene margen_objetivo_cadena_a para cadena A",
    },
    {"cell": "C67", "label": "Contingencia Operativa",   "req": "reglas_negocio.contingencia_operativa.valor"},
    {"cell": "C68", "label": "Contingencia Comercial",   "req": "reglas_negocio.contingencia_comercial.valor"},
    {"cell": "C69", "label": "Mark up",                  "req": "reglas_negocio.markup.valor"},
    {
        "cell": "C70", "label": "Descuento volumen", "req": None,
        "tag": STRUCT_EXT,
        "notes": "Nueva key reglas_negocio.descuento_volumen=0; no existe hoy en req.json",
    },
]

# ── Póliza rows en Panel (col C=activa, D=pct_poliza, E=pct_atr, F=extiende) ─
POLIZA_ROWS: list[dict] = [
    {"row": 38, "req_name": "Póliza de Seriedad",         "tasas_pct": 0.005},
    {"row": 39, "req_name": "Póliza de Cumplimiento",     "tasas_pct": 0.0062},
    {"row": 40, "req_name": "Póliza de Salarios",         "tasas_pct": 0.0119},
    {"row": 41, "req_name": "Poliza de Calidad",          "tasas_pct": 0.0119},
    {"row": 42, "req_name": "Poliza de rc cruzada",       "tasas_pct": 0.0275},
    {"row": 43, "req_name": "poliza de IRF",              "tasas_pct": 0.0275},
    {"row": 44, "req_name": "Póliza de Responsabilidad",  "tasas_pct": 0.0069},
    {"row": 45, "req_name": "Comisión de Administración", "tasas_pct": 0.0118},
    {"row": 46, "req_name": "Otros impuestos",            "tasas_pct": None},
    {"row": 50, "req_name": "Responsabilidad Civil Protección de Datos", "tasas_pct": None},
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _get(d: Any, path: str) -> tuple[Any, bool]:
    """Traverse 'a.b[0].c' dotted+bracket path. Returns (value, found)."""
    parts = re.split(r"[\.\[\]]", path)
    parts = [p for p in parts if p]
    current: Any = d
    for part in parts:
        try:
            if isinstance(current, list):
                current = current[int(part)]
            elif isinstance(current, dict):
                current = current[part]
            else:
                return None, False
        except (KeyError, IndexError, ValueError, TypeError):
            return None, False
    return current, True


def _bool_from_excel(v: Any) -> bool | None:
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in ("si", "sí", "yes", "true", "1"):
            return True
        if s in ("no", "false", "0"):
            return False
    return None


def _num_eq(a: float, b: float) -> bool:
    return abs(a - b) <= TOL_NUM


def _classify_pair(
    v28: Any, req: Any, is_bool: bool = False, allow_suffix: bool = False,
) -> str:
    """Return classification for a matched (v28, req) pair."""
    if v28 is None and req is None:
        return MATCH
    if v28 is None or req is None:
        return VALUE_UPDATE
    if is_bool:
        b28 = _bool_from_excel(v28)
        if b28 is None:
            return UNKNOWN_SRC
        return MATCH if b28 == bool(req) else VALUE_UPDATE
    # Numeric comparison
    if isinstance(v28, (int, float)) and isinstance(req, (int, float)):
        if _num_eq(float(v28), float(req)):
            return MATCH
        scale_hit, _ = is_scale_mismatch(float(req), float(v28))
        if scale_hit:
            return SCALE
        return VALUE_UPDATE
    # String comparison (normalize whitespace + case)
    v28s = str(v28).strip().lower()
    reqs = str(req).strip().lower()
    if allow_suffix and (v28s.endswith(reqs) or reqs.endswith(v28s)):
        return MATCH
    return MATCH if v28s == reqs else VALUE_UPDATE


# ── Section 1: Panel inputs ───────────────────────────────────────────────────

def _map_panel(ws: Any, req: dict) -> list[dict]:
    """Map Panel de Control General cells to request.json paths."""
    rows: list[dict] = []
    for entry in PANEL_MAP:
        cell_ref = entry["cell"]
        label = entry["label"]
        req_path: str | None = entry.get("req")
        tag_override: str | None = entry.get("tag")
        is_bool: bool = entry.get("bool", False)
        notes: str = entry.get("notes", "")

        v28_val = ws[cell_ref].value

        if tag_override:
            # No req.json key → report as STRUCTURE_EXTENSION
            rows.append({
                "hoja_cell": f"Panel!{cell_ref}",
                "label": label,
                "v28": v28_val,
                "req_path": "—",
                "req_val": "—",
                "tipo": tag_override,
                "action": f"Agregar key: {notes}",
                "notes": notes,
            })
            continue

        req_val, found = _get(req, req_path)
        if not found:
            rows.append({
                "hoja_cell": f"Panel!{cell_ref}",
                "label": label,
                "v28": v28_val,
                "req_path": req_path,
                "req_val": "KEY_MISSING",
                "tipo": STRUCT_GAP,
                "action": "Investigar — key no existe",
                "notes": notes,
            })
            continue

        allow_suffix = entry.get("suffix", False)
        tipo = _classify_pair(v28_val, req_val, is_bool, allow_suffix=allow_suffix)

        if is_bool and v28_val is not None:
            v28_display = _bool_from_excel(v28_val)
        elif hasattr(v28_val, "date"):
            # openpyxl returns datetime for date cells; show as ISO string
            v28_display = v28_val.strftime("%Y-%m-%d")
        else:
            v28_display = v28_val

        action = _build_action(tipo, req_path, v28_display, req_val)
        rows.append({
            "hoja_cell": f"Panel!{cell_ref}",
            "label": label,
            "v28": v28_display,
            "req_path": req_path,
            "req_val": req_val,
            "tipo": tipo,
            "action": action,
            "notes": notes,
        })
    return rows


def _build_action(tipo: str, req_path: str, v28: Any, req: Any) -> str:
    if tipo == MATCH:
        return "Ninguna"
    if tipo == VALUE_UPDATE:
        return f"Paso B: cambiar `{req_path}` de `{req}` → `{v28}`"
    if tipo == SCALE:
        return f"Verificar unidad: V2-8={v28} vs req={req} (potencia de 10)"
    return tipo


# ── Section 2: Pólizas ────────────────────────────────────────────────────────

def _map_polizas(ws_panel: Any, req_polizas: list[dict]) -> list[dict]:
    """Compare Panel rows 38-46 + 49-55 (activa, pct, pct_atr, extiende)."""
    rows: list[dict] = []
    req_by_name: dict[str, dict] = {}
    for p in req_polizas:
        # Normalize name for matching
        req_by_name[_norm_name(p["nombre"])] = p

    for prow in POLIZA_ROWS:
        row_num = prow["row"]
        req_name = prow["req_name"]
        tasas_pct: float | None = prow["tasas_pct"]

        v28_name = ws_panel.cell(row=row_num, column=2).value  # col B
        v28_activa = ws_panel.cell(row=row_num, column=3).value  # col C
        v28_pct = ws_panel.cell(row=row_num, column=4).value     # col D
        v28_pct_atr = ws_panel.cell(row=row_num, column=5).value  # col E
        v28_extiende = ws_panel.cell(row=row_num, column=6).value  # col F

        # Find matching req entry: exact first, then prefix/contains fallback
        req_p = (
            req_by_name.get(_norm_name(req_name))
            or req_by_name.get(_norm_name(v28_name or ""))
            or _fuzzy_find(req_by_name, req_name)
            or _fuzzy_find(req_by_name, v28_name or "")
        )

        if req_p is None:
            rows.append({
                "row": row_num,
                "v28_name": v28_name,
                "field": "ALL",
                "v28": f"activa={v28_activa} pct={v28_pct} pct_atr={v28_pct_atr}",
                "req": "NO_MATCH_IN_REQ",
                "tasas": tasas_pct,
                "tipo": STRUCT_GAP,
                "action": "Póliza no encontrada en req.json por nombre",
            })
            continue

        # activa
        tipo_activa = _classify_pair(v28_activa, req_p.get("activa"), is_bool=True)
        # pct_poliza
        tipo_pct, pct_note = _classify_pct_poliza(v28_pct, req_p.get("pct_poliza"), tasas_pct)
        # pct_atribuible
        tipo_pct_atr = _classify_pair(v28_pct_atr, req_p.get("pct_atribuible"))
        # aplica_extension
        tipo_ext = _classify_pair(v28_extiende, req_p.get("aplica_extension"), is_bool=True)

        rows.append({
            "row": row_num,
            "v28_name": v28_name,
            "req_name": req_p["nombre"],
            "v28_activa": _bool_from_excel(v28_activa),
            "req_activa": req_p.get("activa"),
            "tipo_activa": tipo_activa,
            "v28_pct": v28_pct,
            "req_pct": req_p.get("pct_poliza"),
            "tasas_pct": tasas_pct,
            "tipo_pct": tipo_pct,
            "pct_note": pct_note,
            "v28_pct_atr": v28_pct_atr,
            "req_pct_atr": req_p.get("pct_atribuible"),
            "tipo_pct_atr": tipo_pct_atr,
            "v28_extiende": _bool_from_excel(v28_extiende),
            "req_extiende": req_p.get("aplica_extension"),
            "tipo_ext": tipo_ext,
        })
    return rows


def _fuzzy_find(req_by_name: dict[str, dict], name: str) -> dict | None:
    """Find a req entry whose normalized name contains the normalized needle."""
    needle = _norm_name(name)
    if not needle:
        return None
    for key, entry in req_by_name.items():
        if needle in key or key in needle:
            return entry
    return None


def _norm_name(s: str) -> str:
    if not s:
        return ""
    import unicodedata
    # Strip accents, lowercase, collapse whitespace, remove punctuation/parens
    nfkd = unicodedata.normalize("NFKD", s)
    stripped = "".join(c for c in nfkd if not unicodedata.combining(c))
    stripped = re.sub(r"[^a-z0-9 ]", " ", stripped.lower())
    return re.sub(r"\s+", " ", stripped).strip()


def _classify_pct_poliza(
    v28_pct: float | None, req_pct: float | None, tasas_pct: float | None,
) -> tuple[str, str]:
    """
    Returns (tipo, note).
    If Panel (v28_pct) != Tasas (tasas_pct) → EXCEL_LIKELY_BUG.
    Else normal comparison.
    """
    if v28_pct is None:
        return UNKNOWN_SRC, "Panel!D no tiene valor (celda vacía)"
    if tasas_pct is not None and not _num_eq(float(v28_pct), float(tasas_pct)):
        # Internal V2-8 discrepancy: Panel vs Tasas TRM
        return EXCEL_BUG, (
            f"Panel={v28_pct} ≠ Tasas={tasas_pct}; "
            f"motor usa Panel (literal); req.json usa Tasas-aligned value={req_pct}. "
            "DECIDIR en CHECKPOINT A cuál es canónico."
        )
    tipo = _classify_pair(v28_pct, req_pct)
    return tipo, ""


# ── Section 3: Hoja Maestra Escenarios ───────────────────────────────────────

def _extract_scenarios(ws: Any) -> list[dict]:
    """Extract the 3 escenarios from Hoja Maestra Escenarios."""
    scenarios: list[dict] = []
    # Each scenario block starts at A-markers: row 2 (esc1), 50 (esc2), 98 (esc3)
    BLOCKS = [
        ("Escenario 1", 5,  7,  8,  9, 10, 11, 13, 16),
        ("Escenario 2", 53, 55, 56, 57, 58, 59, 61, 64),
        ("Escenario 3", 101, 103, 104, 105, 106, 107, 109, 112),
    ]
    for label, r_esc, r_mod, r_can, r_mod_cob, r_comp_fijo, r_comp_var, r_fte, r_cad_a in BLOCKS:
        def cv(row: int) -> Any:
            return ws.cell(row=row, column=3).value  # col C
        def dv(row: int) -> Any:
            return ws.cell(row=row, column=4).value  # col D (proporcion comp)

        scenarios.append({
            "label": label,
            "escenario": cv(r_esc),
            "modalidad": cv(r_mod),
            "canal": cv(r_can),
            "modelo_cobro": cv(r_mod_cob),
            "componente_fijo": cv(r_comp_fijo),
            "pct_comp_fijo": dv(r_comp_fijo),
            "componente_variable": cv(r_comp_var),
            "pct_comp_var": dv(r_comp_var),
            "fte": cv(r_fte),
            "cadena_a_costo": cv(r_cad_a),
        })
    return scenarios


# ── Section 4: Condiciones Cadena A/B/C ──────────────────────────────────────

def _extract_cadena_a(ws: Any) -> dict:
    """Extract key inputs from Condiciones Cadena A."""
    def cv(r: int, c: int) -> Any:
        return ws.cell(row=r, column=c).value

    perfiles = []
    # Profiles: rows 6-13, columns E (Esc1), F (Esc2), G (Esc3)
    for col, esc_label in [(5, "Escenario SAC Actual"), (6, "Escenario WhatsApp Actual"), (7, "Crecimiento inhouse")]:
        perfiles.append({
            "escenario": esc_label,
            "modalidad": cv(6, col),
            "canal": cv(7, col),
            "fte": cv(9, col),
            "pct_presencia": cv(10, col),
            "estaciones_presenciales": cv(11, col),
            "salario_base": cv(12, col),
            "comision_perfil": cv(13, col),
        })
    return {
        "perfiles": perfiles,
        "nota": "3 escenarios en columnas E/F/G; Esc1=Voz1/130FTE, Esc2=WhatsApp/50FTE, Esc3=Voz2/80FTE",
    }


def _extract_cadena_b(ws: Any) -> dict:
    """Extract OPEX and CAPEX items from Condiciones Cadena B."""
    opex_items: list[dict] = []
    # OPEX header at row 7; data rows 8+
    for row in ws.iter_rows(min_row=8, max_row=42, min_col=2, max_col=8, values_only=False):
        rubro = ws.cell(row=row[0].row, column=2).value
        if rubro is None:
            continue
        opex_items.append({
            "rubro": rubro,
            "modalidad": ws.cell(row=row[0].row, column=3).value,
            "canal": ws.cell(row=row[0].row, column=4).value,
            "tipo_cobro": ws.cell(row=row[0].row, column=6).value,
            "valor": ws.cell(row=row[0].row, column=8).value,
        })
    capex_items: list[dict] = []
    for row in ws.iter_rows(min_row=47, max_row=75, min_col=2, max_col=8, values_only=False):
        rubro = ws.cell(row=row[0].row, column=2).value
        if rubro is None:
            continue
        capex_items.append({
            "rubro": rubro,
            "modalidad": ws.cell(row=row[0].row, column=3).value,
            "canal": ws.cell(row=row[0].row, column=4).value,
            "valor": ws.cell(row=row[0].row, column=6).value,
            "cantidad": ws.cell(row=row[0].row, column=7).value,
            "meses_diferir": ws.cell(row=row[0].row, column=8).value,
        })
    return {
        "opex": [i for i in opex_items if i["rubro"] != "Rubro"],
        "capex": [i for i in capex_items if i["rubro"] != "Rubro"],
        "equipo_soporte_fte": ws.cell(row=79, column=3).value,
    }


def _extract_cadena_c(ws: Any) -> dict:
    """Extract tarifa proveedor and CAPEX from Condiciones Cadena C."""
    tarifa_items: list[dict] = []
    for row in ws.iter_rows(min_row=9, max_row=20, min_col=2, max_col=8, values_only=False):
        proveedor = ws.cell(row=row[0].row, column=2).value
        if proveedor is None:
            continue
        tarifa_items.append({
            "proveedor": proveedor,
            "servicio": ws.cell(row=row[0].row, column=3).value,
            "modalidad": ws.cell(row=row[0].row, column=4).value,
            "canal": ws.cell(row=row[0].row, column=5).value,
            "tipo_cobro": ws.cell(row=row[0].row, column=6).value,
            "valor": ws.cell(row=row[0].row, column=7).value,
            "cantidad": ws.cell(row=row[0].row, column=8).value,
        })
    opex_items: list[dict] = []
    for row in ws.iter_rows(min_row=29, max_row=55, min_col=2, max_col=8, values_only=False):
        desc = ws.cell(row=row[0].row, column=2).value
        if desc is None:
            continue
        opex_items.append({
            "descripcion": desc,
            "modalidad": ws.cell(row=row[0].row, column=3).value,
            "canal": ws.cell(row=row[0].row, column=4).value,
            "valor": ws.cell(row=row[0].row, column=7).value,
            "cantidad": ws.cell(row=row[0].row, column=8).value,
        })
    capex_items: list[dict] = []
    for row in ws.iter_rows(min_row=62, max_row=75, min_col=3, max_col=8, values_only=False):
        modalidad = ws.cell(row=row[0].row, column=3).value
        if modalidad is None:
            continue
        capex_items.append({
            "modalidad": modalidad,
            "canal": ws.cell(row=row[0].row, column=4).value,
            "valor": ws.cell(row=row[0].row, column=6).value,
            "cantidad": ws.cell(row=row[0].row, column=7).value,
            "meses_diferir": ws.cell(row=row[0].row, column=8).value,
        })
    return {
        "tarifas_proveedor": [t for t in tarifa_items if t["proveedor"] != "Proveedor"],
        "opex": [o for o in opex_items if o["descripcion"] not in ("Descripcion",)],
        "capex": [c for c in capex_items if c["modalidad"] not in ("Modalidad",)],
    }


# ── Rendering ─────────────────────────────────────────────────────────────────

def _render(
    panel_rows: list[dict],
    poliza_rows: list[dict],
    scenarios: list[dict],
    req_scenarios: list[dict],
    cadena_a: dict,
    cadena_b: dict,
    cadena_c: dict,
    req: dict,
) -> str:
    from collections import Counter

    tipo_counts: Counter = Counter()
    for r in panel_rows:
        tipo_counts[r["tipo"]] += 1

    poliza_tipos = []
    for p in poliza_rows:
        if "tipo_activa" in p:
            poliza_tipos += [p["tipo_activa"], p["tipo_pct"], p["tipo_pct_atr"], p["tipo_ext"]]
    for t in poliza_tipos:
        tipo_counts[t] += 1
    tipo_counts.pop(MATCH, None)  # remove MATCHes from poliza to keep counting

    total_panel = len(panel_rows)
    total_poliza_rows = len(poliza_rows)

    lines = [
        "# V2-8 Input Mapping — Paso A (CHECKPOINT A)",
        "",
        "> **Estado:** Read-only. `request.json` sin modificar. "
        "Esperando decisión humana antes de Paso B.",
        "",
        "## Resumen ejecutivo",
        "",
        "| Tipo | Panel+Reglas | Pólizas (campo-nivel) |",
        "|------|:---:|:---:|",
    ]
    for tipo in [MATCH, VALUE_UPDATE, SCALE, STRUCT_EXT, STRUCT_GAP, EXCEL_BUG, UNKNOWN_SRC]:
        panel_c = sum(1 for r in panel_rows if r["tipo"] == tipo)
        pol_c = poliza_tipos.count(tipo)
        if panel_c or pol_c:
            lines.append(f"| {tipo} | {panel_c or '—'} | {pol_c or '—'} |")
    lines += [
        "",
        f"- **Total inputs Panel mapeados:** {total_panel}",
        f"- **Total pólizas revisadas:** {total_poliza_rows}",
        "- **Structural gaps (deal-level):** "
        "escenarios_comerciales + condiciones_cadena_a/b/c "
        "(ver §5 — decisión humana requerida)",
        "",
        "---",
        "",
        "## §1 Panel de Control General → datos_operativos / reglas_negocio",
        "",
        "| V2-8 Hoja!Celda | Etiqueta | Valor V2-8 | req.json path | Valor actual | Tipo | Acción propuesta |",
        "|-----------------|----------|-----------|---------------|--------------|------|-----------------|",
    ]
    for r in panel_rows:
        notes_md = f" ⚠ {r['notes']}" if r.get("notes") else ""
        action_md = r.get("action", "—")
        lines.append(
            f"| `{r['hoja_cell']}` | {r['label']} | `{r['v28']}` | "
            f"`{r['req_path']}` | `{r['req_val']}` | **{r['tipo']}** | "
            f"{action_md}{notes_md} |"
        )

    lines += [
        "",
        "---",
        "",
        "## §2 Pólizas (Panel de Control General, filas 38-55)",
        "",
        "> Columnas: C=activa, D=pct\\_poliza, E=pct\\_atribuible, F=aplica\\_extensión",
        "> La columna **Tasas ref.** muestra el valor de `Tasas,TRM,Polizas` para cross-check.",
        "",
        "| Fila | Nombre póliza | V2-8 activa | req activa | Tipo activa | "
        "V2-8 pct | req pct | Tasas ref | Tipo pct | "
        "V2-8 pct_atr | req pct_atr | Tipo pct_atr | "
        "V2-8 ext | req ext | Tipo ext | Nota |",
        "|------|---------------|:-----------:|:----------:|:-----------:|"
        ":-------:|:-------:|:---------:|:--------:|"
        ":----------:|:-----------:|:-----------:|"
        ":-------:|:-------:|:--------:|------|",
    ]
    for p in poliza_rows:
        if "tipo_activa" not in p:
            lines.append(
                f"| {p['row']} | {p.get('v28_name','?')} | — | — | **{p['tipo']}** "
                f"| — | — | — | — | — | — | — | — | — | — | {p.get('action','')} |"
            )
            continue
        pct_note = p.get("pct_note", "")
        note_md = f"⚠ {pct_note}" if pct_note else ""
        lines.append(
            f"| {p['row']} | {p['v28_name']} "
            f"| {p['v28_activa']} | {p['req_activa']} | **{p['tipo_activa']}** "
            f"| {p['v28_pct']} | {p['req_pct']} | {p['tasas_pct']} | **{p['tipo_pct']}** "
            f"| {p['v28_pct_atr']} | {p['req_pct_atr']} | **{p['tipo_pct_atr']}** "
            f"| {p['v28_extiende']} | {p['req_extiende']} | **{p['tipo_ext']}** "
            f"| {note_md} |"
        )

    lines += [
        "",
        "---",
        "",
        "## §3 Fecha de Inicio — nota especial",
        "",
        "| Campo | V2-8 Panel!C10 | request.json | Clasificación |",
        "|-------|---------------|--------------|---------------|",
        "| fecha_inicio | `2026-07-01` | `2026-01-01` | **VALUE_UPDATE** |",
        "",
        "> **Impacto en motor V2-8:** La hoja Visión P&G en V2-8 usa",
        "> `SUMIFS` con gate de fecha para seleccionar el escenario activo.",
        "> Si `fecha_inicio` queda fuera del rango calculado en Hoja Maestra,",
        "> el SUMIFS retornará 0 → ingresos/contingencias = 0 en P&G.",
        "> **Hoja Maestra Escenarios NO tiene columnas de fechas** — el gate",
        "> vive en la fórmula de `Visión P&G` (ver Stage 2 P&G fix).",
        "> Cambiar a `2026-07-01` es necesario para que P&G produzca valores.",
        "",
        "---",
        "",
        "## §4 EXCEL_LIKELY_BUG — Discrepancias internas V2-8 (Panel vs Tasas TRM)",
        "",
        "Las siguientes pólizas tienen valores literales en Panel que **difieren**",
        "del valor de referencia en `Tasas, TRM, Polizas`.",
        "Panel es la fuente que el motor usa (literal input, no fórmula).",
        "Tasas TRM es la tabla de referencia.",
        "",
        "| Póliza | Panel!D (motor usa) | Tasas!B (referencia) | Diferencia | Recomendación |",
        "|--------|--------------------:|---------------------:|:----------:|---------------|",
        "| Póliza de Cumplimiento (D39) | 0.0063 | 0.0062 | +0.0001 | Decidir en CHECKPOINT A: ¿Panel correcto o typo? |",
        "| Poliza de Salarios (D40) | 0.0128 | 0.0119 | +0.0009 | Decidir en CHECKPOINT A: ¿Panel correcto o typo? |",
        "| Poliza de Calidad (D41) | 0.0128 | 0.0119 | +0.0009 | Decidir en CHECKPOINT A: ¿Panel correcto o typo? |",
        "",
        "**ICA adicional:**",
        "",
        "| Fuente | Valor | Nota |",
        "|--------|------:|------|",
        "| Panel!C34 (input deal) | 0.01 | Valor ingresado por usuario para este deal |",
        "| Tasas!B37 (Bogota base tarifa) | 0.00966 | Tarifa base sin sobretasas |",
        "| Tasas!F37 (Bogota total: base+bomberos) | 0.01966 | Incluye 1% bomberos |",
        "| request.json tasa_ica | 0.0097 | Aproximación al base Bogota |",
        "",
        "> ICA Bogota: ninguna fuente coincide exactamente. Confirmar cuál aplicar en CHECKPOINT A.",
        "",
        "---",
        "",
        "## §5 Structural Gaps (deal-level) — decisión humana requerida",
        "",
        "Estos gaps NO pueden resolverse con VALUE_UPDATE. Requieren decisión:",
        "**a)** reemplazar request.json con datos V2-8 (implica regenerar goldens),",
        "**b)** diferir Stage 2 hasta que el deal real del cliente esté en V2-8.",
        "",
        "### §5a escenarios_comerciales",
        "",
        "**V2-8 Hoja Maestra Escenarios (3 escenarios):**",
        "",
        "| # | Escenario | Modalidad | Canal | Modelo cobro | FTE | Comp. Fijo | Comp. Variable | Costo Cad A |",
        "|---|-----------|-----------|-------|--------------|----:|:----------:|:--------------:|------------:|",
    ]
    for s in scenarios:
        lines.append(
            f"| {s['escenario']} | {s['label']} | {s['modalidad']} | {s['canal']} "
            f"| {s['modelo_cobro']} | {s['fte']} "
            f"| {s['componente_fijo']} ({s['pct_comp_fijo']}) "
            f"| {s['componente_variable']} ({s['pct_comp_var']}) "
            f"| {s['cadena_a_costo']:,.0f} |" if isinstance(s['cadena_a_costo'], float)
            else f"| {s['escenario']} | {s['label']} | {s['modalidad']} | {s['canal']} "
            f"| {s['modelo_cobro']} | {s['fte']} "
            f"| {s['componente_fijo']} ({s['pct_comp_fijo']}) "
            f"| {s['componente_variable']} ({s['pct_comp_var']}) "
            f"| {s['cadena_a_costo']} |"
        )

    lines += [
        "",
        "**request.json escenarios_comerciales (actuales):**",
        "",
        "| # | Modalidad | Canal | Modelo cobro | Comp. Fijo | Prop. | Comp. Variable | Prop. |",
        "|---|-----------|-------|--------------|:----------:|------:|:--------------:|------:|",
    ]
    for s in req_scenarios:
        lines.append(
            f"| {s.get('escenario')} | {s.get('modalidad')} | {s.get('canal')} "
            f"| {s.get('modelo_cobro')} "
            f"| {s.get('componente_fijo','')} | {s.get('proporcion_componente_fijo',0)} "
            f"| {s.get('componente_variable','')} | {s.get('proporcion_componente_variable',0)} |"
        )

    lines += [
        "",
        "> **Veredicto:** `REQUEST_STRUCTURE_GAP` (deal-level). Canales completamente distintos",
        "> (V2-8: Voz1+WhatsApp+Voz2; req: WhatsApp+Correo). Modelos cobro distintos.",
        "> Decisión requerida antes de Paso B.",
        "",
        "### §5b Condiciones Cadena A",
        "",
        "**V2-8 tiene 3 perfiles columna (1 por escenario):**",
        "",
        "| Escenario | Modalidad | Canal | FTE | Salario base | Comisión perfil | Estaciones pres. |",
        "|-----------|-----------|-------|----:|-------------:|----------------:|-----------------:|",
    ]
    for p in cadena_a.get("perfiles", []):
        lines.append(
            f"| {p['escenario']} | {p['modalidad']} | {p['canal']} "
            f"| {p['fte']} | {p['salario_base']:,.0f} | {p['comision_perfil']:,.0f} "
            f"| {p['estaciones_presenciales']} |"
        )

    req_cad_a_perfiles = req.get("condiciones_cadena_a", {}).get("perfiles", [])
    lines += [
        "",
        "**request.json condiciones_cadena_a.perfiles (actuales):**",
        "",
        "| Perfil | Modalidad | Canal | FTE |",
        "|--------|-----------|-------|----:|",
    ]
    for p in req_cad_a_perfiles:
        lines.append(
            f"| {p.get('nombre','?')} | {p.get('modalidad','?')} "
            f"| {p.get('canal','?')} | {p.get('fte','?')} |"
        )

    lines += [
        "",
        "> **Veredicto:** `REQUEST_STRUCTURE_GAP` (deal-level). V2-8 tiene 3 escenarios por columna;",
        "> req.json tiene 3 perfiles por fila con canales distintos (WhatsApp/Correo/WebChat).",
        "> Estructura incompatible; no parcheable por VALUE_UPDATE solo.",
        "",
        "### §5c Condiciones Cadena B",
        "",
        "**V2-8 OPEX items (Condiciones Cadena B):**",
        "",
        "| Rubro | Modalidad | Canal | Tipo cobro | Valor |",
        "|-------|-----------|-------|------------|------:|",
    ]
    for item in cadena_b.get("opex", []):
        if item["rubro"] and item["rubro"] not in ("Rubro",):
            lines.append(
                f"| {item['rubro']} | {item.get('modalidad','')} "
                f"| {item.get('canal','')} | {item.get('tipo_cobro','')} "
                f"| {item.get('valor','')} |"
            )

    lines += [
        "",
        f"**V2-8 Equipo Soporte FTE:** {cadena_b.get('equipo_soporte_fte')}",
        "",
        "**V2-8 CAPEX items:**",
        "",
        "| Rubro | Modalidad | Canal | Valor | Cantidad | Meses diferir |",
        "|-------|-----------|-------|------:|---------:|--------------:|",
    ]
    for item in cadena_b.get("capex", []):
        lines.append(
            f"| {item.get('rubro','')} | {item.get('modalidad','')} "
            f"| {item.get('canal','')} | {item.get('valor','')} "
            f"| {item.get('cantidad','')} | {item.get('meses_diferir','')} |"
        )

    lines += [
        "",
        "> **Veredicto:** `REQUEST_STRUCTURE_GAP`. V2-8 Cadena B tiene OPEX para canal Voz2;",
        "> req.json tiene diferentes rubros/canales.",
        "",
        "### §5d Condiciones Cadena C",
        "",
        "**V2-8 tarifas proveedor:**",
        "",
        "| Proveedor | Servicio | Modalidad | Canal | Tipo cobro | Valor | Cantidad |",
        "|-----------|----------|-----------|-------|------------|------:|---------:|",
    ]
    for t in cadena_c.get("tarifas_proveedor", []):
        lines.append(
            f"| {t.get('proveedor','')} | {t.get('servicio','')} "
            f"| {t.get('modalidad','')} | {t.get('canal','')} "
            f"| {t.get('tipo_cobro','')} | {t.get('valor','')} | {t.get('cantidad','')} |"
        )

    lines += [
        "",
        "**V2-8 OPEX Cadena C:**",
        "",
        "| Descripción | Modalidad | Canal | Valor | Cantidad |",
        "|-------------|-----------|-------|------:|---------:|",
    ]
    for o in cadena_c.get("opex", []):
        lines.append(
            f"| {o.get('descripcion','')} | {o.get('modalidad','')} "
            f"| {o.get('canal','')} | {o.get('valor','')} | {o.get('cantidad','')} |"
        )

    lines += [
        "",
        "> **Veredicto:** `REQUEST_STRUCTURE_GAP`. V2-8 Cadena C usa Accenture/Nexa AI/Voz1;",
        "> req.json condiciones_cadena_c actualmente vacío (tarifa_proveedor_canal.items=[]).",
        "> Este es un cambio significativo de deal.",
        "",
        "---",
        "",
        "## §6 STRUCTURE_EXTENSION — keys nuevas dentro de objetos existentes",
        "",
        "Los siguientes inputs V2-8 no tienen key en req.json pero pueden **agregarse**",
        "dentro de un objeto top-level existente (sin nueva key top-level, con aprobación):",
        "",
        "| V2-8 Cell | Label | Valor V2-8 | Propuesta key nueva | Objeto existente |",
        "|-----------|-------|-----------|---------------------|-----------------|",
    ]
    for r in panel_rows:
        if r["tipo"] == STRUCT_EXT:
            lines.append(
                f"| `{r['hoja_cell']}` | {r['label']} | `{r['v28']}` "
                f"| {r['notes']} | (top-level existente) |"
            )

    lines += [
        "",
        "---",
        "",
        "## §7 Confirmaciones (sin cambio necesario)",
        "",
        "- `Riesgo`: NO tiene celdas de input del deal. Solo parametrización estática "
        "(pesos de calificación de riesgo operativo/cliente). Sin acción.",
        "- `Pólizas - Costo Financiacion`: NO tiene celdas de input. "
        "Hoja de cálculo/reporte; referencia Panel y Costos Totales. Sin acción.",
        "- `Tasas, TRM, Polizas`: Tabla de referencia (IPC, SMLV, ICA por ciudad). "
        "Bogota: base=Tasas!B37=0.00966, total con sobretasas=Tasas!F37=0.01966. "
        "Estos son parámetros de parametrización, no inputs del deal.",
        "",
        "---",
        "",
        "## CHECKPOINT A — Decisiones requeridas antes de Paso B",
        "",
        "1. **escenarios_comerciales + condiciones_cadena_a/b/c** (§5):",
        "   - Opción a: reemplazar request.json con deal V2-8 (SAC/METROCUADRADO/Voz1)",
        "     → regeneración total de 63 goldens necesaria",
        "   - Opción b: diferir hasta que el deal real esté en V2-8",
        "",
        "2. **EXCEL_LIKELY_BUG pólizas** (§4):",
        "   - Cumplimiento D39=0.0063 vs Tasas B22=0.0062 → ¿cuál es canónico?",
        "   - Salarios D40=0.0128 vs Tasas B23=0.0119 → ¿cuál es canónico?",
        "   - Calidad D41=0.0128 vs Tasas B24=0.0119 → ¿cuál es canónico?",
        "",
        "3. **ICA discrepancia** (§4): Panel!C34=0.01 vs Tasas Bogota=0.00966/0.01966 vs req=0.0097",
        "",
        "4. **STRUCTURE_EXTENSION** (§6): Aprobar key por key antes de Paso B.",
        "",
        "5. **VALUE_UPDATE** (§1): Una vez resueltas las decisiones anteriores,",
        "   los VALUE_UPDATE restantes pueden aplicarse en Paso B.",
        "",
    ]
    return "\n".join(lines)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    import sys

    # Sanity check: backup must exist
    bak = REQUEST_PATH.parent / (REQUEST_PATH.name + ".pre_v28_parity.bak")
    if not bak.exists():
        print(f"ERROR: backup not found at {bak}", file=sys.stderr)
        print("Crear con: cp request/request.json request/request.json.pre_v28_parity.bak",
              file=sys.stderr)
        return 1

    print(f"Opening {EXCEL_V28_PATH.name} (data_only=True)…")
    wb = openpyxl.load_workbook(str(EXCEL_V28_PATH), data_only=True)

    ws_panel = wb["Panel de Control General"]
    ws_tasas = wb["Tasas, TRM, Polizas"]
    ws_hoja = wb["Hoja Maestra Escenarios"]
    ws_cad_a = wb["Condiciones Cadena A"]
    ws_cad_b = wb["Condiciones Cadena B"]
    ws_cad_c = wb["Condiciones Cadena C"]

    req: dict = json.loads(REQUEST_PATH.read_text())

    # Verify request.json is unmodified (compare with backup)
    bak_data = json.loads(bak.read_text())
    if req != bak_data:
        print("WARNING: request.json differs from backup — Paso A should be read-only!",
              file=sys.stderr)
        return 1

    print("Mapping Panel de Control General…")
    panel_rows = _map_panel(ws_panel, req)

    print("Mapping pólizas…")
    poliza_rows = _map_polizas(ws_panel, req.get("polizas", []))

    print("Extracting Hoja Maestra Escenarios…")
    scenarios = _extract_scenarios(ws_hoja)

    print("Extracting Condiciones Cadena A/B/C…")
    cadena_a = _extract_cadena_a(ws_cad_a)
    cadena_b = _extract_cadena_b(ws_cad_b)
    cadena_c = _extract_cadena_c(ws_cad_c)

    print("Rendering markdown…")
    md = _render(
        panel_rows,
        poliza_rows,
        scenarios,
        req.get("escenarios_comerciales", []),
        cadena_a,
        cadena_b,
        cadena_c,
        req,
    )

    OUT_MD.write_text(md, encoding="utf-8")
    print(f"→ {OUT_MD.relative_to(BACKEND_ROOT)}")

    # Print summary
    from collections import Counter
    counts: Counter = Counter(r["tipo"] for r in panel_rows)
    print("\nPanel summary:")
    for tipo, n in sorted(counts.items()):
        print(f"  {tipo}: {n}")
    print(f"\nPólizas revisadas: {len(poliza_rows)}")
    print("\nCHECKPOINT A — mapping completo. Esperar decisión humana.")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
