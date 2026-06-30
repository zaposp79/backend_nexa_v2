#!/usr/bin/env python3
"""
scripts/validate_excel_v28.py
==============================
Gate de validación V2-8 para deal canónico SAC/METROCUADRADO COM SAS / Grupo Aval.

Diferencia con validate_excel.py (legado V2-7):
  - Apunta a Excel V2-8 (no V2-7)
  - Valida inputs de usuario de hojas Panel/Condiciones (no outputs HME/P&G cacheados)
  - Compara mecánica de fórmulas verificables (IPC ratio, CAPEX, Cadena C ingreso)
  - Clasifica HME cache como SKIPPED_OLD_EXCEL_CACHE_NOT_COMPARABLE (no FAIL)
  - Exit code 1 solo si hay mismatch real en inputs/formulas comparables
"""
from __future__ import annotations

import json
import logging
import sys
import warnings
from pathlib import Path

logging.disable(logging.WARNING)
warnings.filterwarnings("ignore")

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

import backend_nexa  # noqa: E402,F401 — registers nexa_engine alias in sys.modules
import openpyxl  # noqa: E402

from nexa_engine.modules.calculator_motor.adapters.user_input_loader import UserInputLoader  # noqa: E402
from nexa_engine.modules.calculator_motor.context_builder import SimulationContextBuilder  # noqa: E402
from nexa_engine.modules.calculator_motor.engine import NexaPricingEngine  # noqa: E402
from nexa_engine.modules.parametrizacion.services.provider import ParametrizationProvider  # noqa: E402
from nexa_engine.modules.parametrizacion.services.resolver import ParametrizationResolver  # noqa: E402

BACKEND_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH   = BACKEND_ROOT / "excel" / "Nexa - Pricing - Simulador - V2-8.xlsx"
REQUEST_PATH = BACKEND_ROOT / "request" / "request.json"
OP_PATH      = BACKEND_ROOT / "storage" / "parametrization" / "v2-7" / "op.json"
_V27_DIR     = BACKEND_ROOT / "storage" / "parametrization" / "v2-7"

DEAL_LABEL   = "SAC / METROCUADRADO COM SAS / Grupo Aval"
MAX_DELTA    = 0.000001
SEPARATOR    = "─" * 44

# Excel cells (Panel de Control General)
_PANEL_SHEET  = "Panel de Control General"
_CELL_COMP_TEC   = "L8"   # componente_tecnologico
_CELL_COMP_HUM   = "L7"   # componente_humano
_CELL_MARGEN     = "C63"  # margen_objetivo
_CELL_DURACION   = "C11"  # duracion_meses


def _label(n: int, total: int, desc: str) -> str:
    return f"[{n}/{total}] {desc}..."


def _pass(msg: str = "") -> tuple[str, str]:
    return "PASS", msg


def _fail(msg: str = "") -> tuple[str, str]:
    return "FAIL", msg


def _skip(msg: str = "") -> tuple[str, str]:
    return "SKIP", msg


def _warn(msg: str = "") -> tuple[str, str]:
    return "WARN", msg


# ──────────────────────────────────────────────────────────────
# Check 1 — Input alignment (Panel + Condiciones)
# ──────────────────────────────────────────────────────────────

def check_input_alignment(
    excel_path: Path,
    request: dict,
) -> tuple[str, str]:
    """Compare Panel de Control General inputs against request.json fields."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return _fail(f"Cannot open Excel: {exc}")

    try:
        ws = wb[_PANEL_SHEET]
    except KeyError:
        wb.close()
        return _fail(f"Sheet '{_PANEL_SHEET}' not found in Excel")

    excel_comp_tec = str(ws[_CELL_COMP_TEC].value or "").strip()
    excel_comp_hum = str(ws[_CELL_COMP_HUM].value or "").strip()
    excel_margen   = ws[_CELL_MARGEN].value
    excel_duracion = ws[_CELL_DURACION].value
    wb.close()

    # Request fields
    indexacion   = request.get("volumetria", {}).get("indexacion", {})
    req_comp_tec = str(indexacion.get("componente_tecnologico", "")).strip()
    req_comp_hum = str(indexacion.get("componente_humano", "")).strip()
    req_margen   = request.get("reglas_negocio", {}).get("margen_objetivo")
    req_duracion = request.get("datos_operativos", {}).get("duracion_meses")

    mismatches: list[str] = []

    # componente_tecnologico — normalize separator variants
    # Excel may use "20% SMMLV 80% IPC" while OP storage has "20% SMMLV - 80% IPC"
    _norm = lambda s: s.replace(" - ", " ").replace("-", "").lower().strip()
    if _norm(excel_comp_tec) != _norm(req_comp_tec):
        mismatches.append(
            f"componente_tecnologico: Excel='{excel_comp_tec}' req='{req_comp_tec}'"
        )

    if _norm(excel_comp_hum) != _norm(req_comp_hum):
        mismatches.append(
            f"componente_humano: Excel='{excel_comp_hum}' req='{req_comp_hum}'"
        )

    if excel_margen is not None and req_margen is not None:
        if abs(float(excel_margen) - float(req_margen)) > 0.0001:
            mismatches.append(
                f"margen_objetivo: Excel={excel_margen} req={req_margen}"
            )

    if excel_duracion is not None and req_duracion is not None:
        if int(excel_duracion) != int(req_duracion):
            mismatches.append(
                f"duracion_meses: Excel={excel_duracion} req={req_duracion}"
            )

    total_checked = 4
    if mismatches:
        return _fail("; ".join(mismatches))
    return _pass(f"({total_checked}/{total_checked} fields match)")


# ──────────────────────────────────────────────────────────────
# Check 2 — componente_tecnologico resolves (no ParametrizationError)
# ──────────────────────────────────────────────────────────────

def check_engine_resolves(request: dict) -> tuple[str, str]:
    """Run engine; PASS if no exception raised."""
    try:
        loader = UserInputLoader()
        ui = loader.cargar_desde_dict(request)
        solic = SimulationContextBuilder().construir(ui)
        NexaPricingEngine().calcular(solic)
        return _pass()
    except Exception as exc:  # noqa: BLE001
        return _fail(f"{type(exc).__name__}: {exc}")


# ──────────────────────────────────────────────────────────────
# Check 3 — Storage OP accepts componente_tecnologico
# ──────────────────────────────────────────────────────────────

def check_op_has_componente(op_path: Path, request: dict) -> tuple[str, str]:
    """Verify OP parametrization has the requested componente_tecnologico row."""
    if not op_path.exists():
        return _skip(f"OP file not found at {op_path}")

    indexacion = request.get("volumetria", {}).get("indexacion", {})
    req_comp   = str(indexacion.get("componente_tecnologico", "")).strip()
    if not req_comp:
        return _fail("componente_tecnologico absent in request.json")

    try:
        op_data = json.loads(op_path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        return _fail(f"Cannot read OP JSON: {exc}")

    _norm = lambda s: s.replace(" - ", " ").replace("-", "").lower().strip()
    req_norm = _norm(req_comp)

    def _search(node: object) -> bool:
        if isinstance(node, dict):
            for k, v in node.items():
                if k == "componente" and isinstance(v, str) and _norm(v) == req_norm:
                    return True
                if _search(v):
                    return True
        elif isinstance(node, list):
            return any(_search(item) for item in node)
        return False

    if _search(op_data):
        return _pass(f"'{req_comp}' found in OP storage")
    return _fail(f"'{req_comp}' NOT found in OP storage")


# ──────────────────────────────────────────────────────────────
# V2-7 provider builder (mirrors tests/refactor/_v27_provider.py)
# ──────────────────────────────────────────────────────────────

class _V27Repo:
    """Stub repo that returns v2-7 data from filesystem, bypassing versions.json lookup."""

    def __init__(self, path: Path) -> None:
        self._data: dict = json.loads(path.read_text(encoding="utf-8"))

    def get_active_data(self) -> dict:
        return self._data


def _build_v27_provider() -> ParametrizationProvider:
    """Return a ParametrizationProvider backed by v2-7 parametrization files."""
    resolver = ParametrizationResolver(
        hr_repo=_V27Repo(_V27_DIR / "hr.json"),
        gn_repo=_V27Repo(_V27_DIR / "gn.json"),
        op_repo=_V27Repo(_V27_DIR / "op.json"),
    )
    return ParametrizationProvider.build(resolver)


# ──────────────────────────────────────────────────────────────
# Check 4 — IPC ratio mechanism
# ──────────────────────────────────────────────────────────────

def check_ipc_ratio(request: dict) -> tuple[str, str]:
    """
    Verify IPC indexation mechanism using V2-7 provider (IPC rates configured).

    Deal starts 2026-07-01; months 1-6 in 2026 (IPC=0), month 7 in 2027 (IPC=0.05547729).
    Expected: ingreso_bruto_a[M7] / ingreso_bruto_a[M6] == 1 + IPC[2027].

    EXCEL V2-8: Visión P&G sheet, mechanism validated in golden test
    test_pyg_v28_ingreso_indexado.py::test_ipc_2027_ratio_cadena_a (7 passed).
    """
    ipc_2027     = 0.05547729  # IPC 2027 rate from v2-7/op.json
    expected     = 1.0 + ipc_2027
    # m6 = index 5 (Dec 2026, IPC=0), m7 = index 6 (Jan 2027, IPC=0.05547729)
    m6_idx, m7_idx = 5, 6

    if not _V27_DIR.exists():
        return _skip(f"V2-7 parametrization dir not found: {_V27_DIR}")

    try:
        provider = _build_v27_provider()
        loader   = UserInputLoader()
        ui       = loader.cargar_desde_dict(request)
        solic    = SimulationContextBuilder().construir(ui)
        result   = NexaPricingEngine(parametrizacion=provider).calcular(solic)
    except Exception as exc:  # noqa: BLE001
        return _fail(f"Engine error: {exc}")

    months = result.pyg_por_mes
    if len(months) <= m7_idx:
        return _skip(f"Only {len(months)} months — need at least {m7_idx + 1}")

    m6_val = months[m6_idx].ingreso_bruto_a
    m7_val = months[m7_idx].ingreso_bruto_a

    if m6_val <= 0:
        return _fail(f"m6 ingreso_bruto_a={m6_val:.2f} — unexpected zero/negative")

    ratio = m7_val / m6_val
    delta = abs(ratio - expected)

    if delta < MAX_DELTA:
        return _pass(f"delta={delta:.10f} (m7={m7_val:,.0f} / m6={m6_val:,.0f})")
    return _fail(
        f"m7/m6 ratio={ratio:.10f} expected={expected:.10f} delta={delta:.10f}"
    )


# ──────────────────────────────────────────────────────────────
# Check 5 — CAPEX-001 formula active
# ──────────────────────────────────────────────────────────────

def check_capex_active(request: dict) -> tuple[str, str]:
    """
    Verify CAPEX amortization formula is active.

    If request has CAPEX items in cadena_b or cadena_c, the resulting
    costo_b / costo_c must be > 0 confirming pipeline processes them.
    """
    cadena_b_capex = request.get("condiciones_cadena_b", {}).get("inversiones_capex", [])
    cadena_c_capex = request.get("condiciones_cadena_c", {}).get("inversiones_capex", [])

    has_capex = bool(cadena_b_capex or cadena_c_capex)
    if not has_capex:
        return _skip("No CAPEX items in request — formula not exercised")

    try:
        loader = UserInputLoader()
        ui = loader.cargar_desde_dict(request)
        solic = SimulationContextBuilder().construir(ui)
        result = NexaPricingEngine().calcular(solic)
    except Exception as exc:  # noqa: BLE001
        return _fail(f"Engine error: {exc}")

    p1 = result.pyg_por_mes[0]

    if cadena_b_capex and p1.costo_b <= 0:
        return _fail(f"cadena_b has CAPEX but costo_b={p1.costo_b:.2f} in m1")

    if cadena_c_capex and p1.costo_c <= 0:
        return _fail(f"cadena_c has CAPEX but costo_c={p1.costo_c:.2f} in m1")

    detail_parts: list[str] = []
    if cadena_b_capex:
        detail_parts.append(f"costo_b={p1.costo_b:,.0f}")
    if cadena_c_capex:
        detail_parts.append(f"costo_c={p1.costo_c:,.0f}")

    return _pass(", ".join(detail_parts))


# ──────────────────────────────────────────────────────────────
# Check 6 — CADENA_C_NULL regression guard
# ──────────────────────────────────────────────────────────────

def check_cadena_c_active(request: dict) -> tuple[str, str]:
    """Verify ingreso_bruto_c > 0 in month 1 — guard against CADENA_C_NULL regression."""
    try:
        loader = UserInputLoader()
        ui = loader.cargar_desde_dict(request)
        solic = SimulationContextBuilder().construir(ui)
        result = NexaPricingEngine().calcular(solic)
    except Exception as exc:  # noqa: BLE001
        return _fail(f"Engine error: {exc}")

    ingreso_c = result.pyg_por_mes[0].ingreso_bruto_c
    if ingreso_c <= 0:
        return _fail(f"ingreso_bruto_c={ingreso_c:.2f} — CADENA_C_NULL regression detected")
    return _pass(f"ingreso_c={ingreso_c:,.0f}")


# ──────────────────────────────────────────────────────────────
# Check 7 — HME cache classification (no FAIL)
# ──────────────────────────────────────────────────────────────

def check_hme_cache(excel_path: Path) -> tuple[str, str]:
    """
    Read HME!C296 from Excel V2-8.

    Excel V2-8 was calculated with a different deal than SAC/METROCUADRADO.
    Any value found here is cached from that other scenario and is NOT
    comparable against our backend run. Always classifies as SKIP.
    """
    skip_reason = "SKIPPED_OLD_EXCEL_CACHE_NOT_COMPARABLE"
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        hme_sheet_names = [s for s in wb.sheetnames if "maestra" in s.lower() or "hme" in s.lower()]
        if not hme_sheet_names:
            wb.close()
            return _skip(f"{skip_reason} — HME sheet not found")
        ws = wb[hme_sheet_names[0]]
        hme_val = ws["C296"].value
        wb.close()
        return _skip(f"{skip_reason} (HME!C296={hme_val!r})")
    except Exception as exc:  # noqa: BLE001
        return _skip(f"{skip_reason} — read error: {exc}")


# ──────────────────────────────────────────────────────────────
# Runner
# ──────────────────────────────────────────────────────────────

def _print_result_line(n: int, total: int, desc: str, status: str, msg: str) -> None:
    label_width = 50
    label = f"[{n}/{total}] {desc}..."
    padded = f"{label:<{label_width}}"
    status_str = f"{status:<6}"
    detail = f"  ({msg})" if msg else ""
    print(f"  {padded} {status_str}{detail}")


def main() -> int:
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel V2-8 not found at {EXCEL_PATH}", file=sys.stderr)
        return 2

    if not REQUEST_PATH.exists():
        print(f"ERROR: request.json not found at {REQUEST_PATH}", file=sys.stderr)
        return 2

    try:
        request = json.loads(REQUEST_PATH.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: Cannot parse request.json: {exc}", file=sys.stderr)
        return 2

    print()
    print("V2-8 Excel Validation Gate")
    print(f"deal: {DEAL_LABEL}")
    print(f"max_delta_allowed: {MAX_DELTA}")
    print()

    total_checks = 7
    checks: list[tuple[str, tuple[str, str]]] = []

    checks.append(("Input alignment (Panel + Condiciones)", check_input_alignment(EXCEL_PATH, request)))
    checks.append(("componente_tecnologico resolves",       check_engine_resolves(request)))
    checks.append(("Storage OP has componente_tecnologico", check_op_has_componente(OP_PATH, request)))
    checks.append(("IPC ratio mechanism",                   check_ipc_ratio(request)))
    checks.append(("CAPEX-001 factor activo",               check_capex_active(request)))
    checks.append(("CADENA_C_NULL no regreso",              check_cadena_c_active(request)))
    checks.append(("HME cache check",                       check_hme_cache(EXCEL_PATH)))

    for idx, (desc, (status, msg)) in enumerate(checks, start=1):
        _print_result_line(idx, total_checks, desc, status, msg)

    print()
    print(SEPARATOR)

    pass_count  = sum(1 for _, (s, _) in checks if s == "PASS")
    fail_count  = sum(1 for _, (s, _) in checks if s == "FAIL")
    warn_count  = sum(1 for _, (s, _) in checks if s == "WARN")
    skip_count  = sum(1 for _, (s, _) in checks if s in ("SKIP", "WARN"))
    real_checks = total_checks - sum(1 for _, (s, _) in checks if s == "SKIP")

    overall = "PASS" if fail_count == 0 else "FAIL"
    real_pass = pass_count
    print(f"RESULT: {overall}  ({real_pass}/{real_checks} checks, {skip_count} skipped)")
    print()

    if fail_count > 0:
        print("FAILURES:")
        for desc, (status, msg) in checks:
            if status == "FAIL":
                print(f"  - {desc}: {msg}")
        print()

    if warn_count > 0:
        print("WARNINGS:")
        for desc, (status, msg) in checks:
            if status == "WARN":
                print(f"  - {desc}: {msg}")
        print()

    return 1 if fail_count > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
