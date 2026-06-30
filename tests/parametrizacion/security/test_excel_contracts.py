"""Tests for contract-based sheet/header validation and full upload pipeline.

Tests are organised in three sections:
  1. Reader-level contract validation (no HTTP layer).
  2. Value normalizer injection detection.
  3. Full HTTP upload pipeline (positive and negative).
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from nexa_engine.modules.parametrizacion.shared.helpers.excel_reader import read_excel_sheets
from nexa_engine.modules.parametrizacion.shared.helpers.value_normalizer import normalize_all_sheets_values
from nexa_engine.modules.shared.exceptions import UploadError, ValidationError

BACKEND_ROOT = Path(__file__).resolve().parents[3]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _workbook_bytes(sheets: dict) -> bytes:
    """Build xlsx bytes from {sheet_name: (headers_list, rows_list)}."""
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(list(headers))
        for row in rows:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _valid_hr_workbook() -> bytes:
    """Minimal HR workbook that satisfies all required sheet + header contracts."""
    return _workbook_bytes({
        "HR-LV": (
            ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
            [["Empleado", "Agente", "Cobranzas", "Cesantías", "Salud", "Recargo festivo"]],
        ),
        "HR-SalarioBasico": (["Servicio", "Valor"], [["Cobranzas", 1423000]]),
        "HR-Nomina": (["Tipo", "Rol", "Salario", None, None], [["Operativo", "Agente", 1423000, None, None]]),
        "HR-Recargos": (["Recargo", "Valor"], [["Nocturno", 0.35]]),
        "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["Salud", 0.085]]),
        "HR-Prestaciones": (["Prestaciones", "Valor"], [["Cesantias", 0.0833]]),
        "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Supervisor", "Cobranzas", "Operativo", 20]]),
    })


def _valid_gn_workbook() -> bytes:
    headers = [
        "Ciudad", "Localidad", "Servicio", "CategoriaServicio", "CentroCosto",
        "Componente", "Poliza", "ComponenteFijo", "HardwareSoftware", "PeriodoPago",
        "Cadena", "ComponenteVariable", "ModeloCombro", "Modalidad", "ReglaNegocio",
        "Canal", "Metrica", "Cliente", "TipoCobro", "TipoCliente", "Rubro", "UnidadMedida",
        "Divisa",
    ]
    return _workbook_bytes({
        "GN-LV": (headers, [["Bogota", "Bogota", "Cobranzas", "Cuentas", "CC01",
                             "Op", "Pol1", "Fijo", "PC", "Mensual",
                             "A", "Var1", "Fijo", "Inbound", "R1",
                             "Digital", "M1", "BancaMia", "TC1", "TipoA", "Rubro1", "Und",
                             "COP"]]),
    })


def _valid_op_workbook() -> bytes:
    return _workbook_bytes({
        "OP-OPEXFijo": (["OPEXItem", "Valor"], [["Internet", 45000]]),
    })


# ===========================================================================
# 1. Reader-level contract validation
# ===========================================================================

class TestReaderContractSheets:

    def test_rejects_missing_required_sheet(self):
        """HR-LV is required but absent."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cobranzas", 1423000]]),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        assert any("HR-LV" in e for e in exc_info.value.errors)

    def test_rejects_unauthorized_extra_sheet(self):
        """HR-Desconocida is not in the contract."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["Empleado", "Agente", "Cobranzas", "Ces", "Salud", "RF"]],
            ),
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cobranzas", 1]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario"], [["Op", "Ag", 1]]),
            "HR-Recargos": (["Recargo", "Valor"], [["N", 0.3]]),
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["S", 0.08]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.08]]),
            "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "Cob", "Op", 20]]),
            "HR-Desconocida": (["Col1"], [["val"]]),  # not authorized
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        assert any("HR-Desconocida" in e for e in exc_info.value.errors)

    def test_rejects_renamed_required_sheet(self):
        """HR-LV renamed to HR-LV-Mod."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV-Mod": (["Tipo", "Rol"], [["E", "A"]]),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        # Both: required HR-LV missing AND unauthorized HR-LV-Mod
        errors_combined = " ".join(exc_info.value.errors)
        assert "HR-LV" in errors_combined

    def test_rejects_too_many_sheets(self, monkeypatch):
        import nexa_engine.modules.parametrizacion.shared.helpers.excel_reader as reader
        monkeypatch.setattr(reader, "MAX_EXCEL_SHEETS", 2)
        wb = openpyxl.Workbook()
        for i in range(4):
            wb.create_sheet(f"HR-Sheet{i}")
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        with pytest.raises(UploadError) as exc_info:
            read_excel_sheets(buf.getvalue(), "HR-")
        assert exc_info.value.code == "EXCEL_LIMIT_EXCEEDED"


class TestReaderContractHeaders:

    def test_rejects_missing_header(self):
        """HR-LV is present but missing 'Recargo' column."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales"],  # missing Recargo
                [["E", "A", "Cob", "Ces", "Sal"]],
            ),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        errors = " ".join(exc_info.value.errors)
        assert "HR-LV" in errors

    def test_rejects_extra_named_header(self):
        """HR-SalarioBasico has an extra named column 'Extra'."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
            "HR-SalarioBasico": (["Servicio", "Valor", "Extra"], [["Cob", 1, "x"]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario"], [["Op", "Ag", 1]]),
            "HR-Recargos": (["Recargo", "Valor"], [["N", 0.3]]),
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["S", 0.08]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.08]]),
            "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "C", "Op", 20]]),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        assert any("HR-SalarioBasico" in e for e in exc_info.value.errors)

    def test_rejects_wrong_accent_in_header(self):
        """'Prestaciones' written without accent but contract requires exact match."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                # 'Prestaciones' is correct here, but let's break SS&Parafiscales
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SSParafiscales", "Recargo"],
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        errors = " ".join(exc_info.value.errors)
        assert "SS&Parafiscales" in errors or "SSParafiscales" in errors

    def test_rejects_wrong_capitalization_in_header(self):
        """'servicio' (lowercase) instead of 'Servicio'."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        errors = " ".join(exc_info.value.errors)
        assert "Rol" in errors or "rol" in errors

    def test_rejects_duplicate_header(self):
        """Duplicate column name caught at exact positional comparison.

        Contract expects: Tipo, Rol, Servicio, Prestaciones, SS&Parafiscales, Recargo
        File has:         Tipo, Tipo, Servicio, Prestaciones, SS&Parafiscales, Recargo
                               ^^^^
        Position 2 receives "Tipo" but expects "Rol" → ValidationError.
        """
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Tipo", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["Empleado", "Empleado2", "Cob", "Ces", "Sal", "RF"]],
            ),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        # Position 2: "Tipo" ≠ expected "Rol"
        errors = " ".join(exc_info.value.errors)
        assert "Rol" in errors or "Tipo" in errors

    def test_rejects_unnamed_style_header(self):
        """Header 'Unnamed: 0' is not in the contract — rejected as wrong header.

        openpyxl returns non-empty strings for named columns; 'Unnamed: 0' is
        a pandas-style artifact that could appear if the sheet was produced by
        a data-engineering tool.  It must be rejected, not silently skipped.
        """
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Unnamed: 0", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["val0", "Agente", "Cob", "Ces", "Sal", "RF"]],
            ),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        # "Unnamed: 0" ≠ "Tipo" at position 1
        errors = " ".join(exc_info.value.errors)
        assert "Tipo" in errors or "Unnamed" in errors

    def test_rejects_internal_space_change_in_header(self):
        """'SS & Parafiscales' (spaces around &) ≠ 'SS&Parafiscales' (production).

        Internal whitespace changes are NOT tolerated; only leading/trailing
        spaces are stripped by the contract reader.
        """
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS & Parafiscales", "Recargo"],
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
        })
        with pytest.raises(ValidationError) as exc_info:
            read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        errors = " ".join(exc_info.value.errors)
        assert "SS&Parafiscales" in errors or "SS & Parafiscales" in errors

    def test_accepts_trailing_space_in_header(self):
        """'Cargo ' (trailing space, as in production) should be accepted.

        The contract defines 'Cargo' (stripped) and the reader strips
        whitespace from raw headers before comparing.
        """
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cob", 1]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario"], [["Op", "Ag", 1]]),
            "HR-Recargos": (["Recargo", "Valor"], [["N", 0.3]]),
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["S", 0.08]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.08]]),
            "HR-Ratios": (["Cargo ", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "C", "Op", 20]]),
        })
        # Must not raise — 'Cargo ' stripped to 'Cargo' matches contract
        result = read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        assert "HR-Ratios" in result

    def test_accepts_trailing_none_columns_when_allowed(self):
        """HR-Nomina allows trailing unnamed columns (production has 2)."""
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cob", 1]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario", None, None], [["Op", "Ag", 1, None, None]]),
            "HR-Recargos": (["Recargo", "Valor"], [["N", 0.3]]),
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["S", 0.08]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.08]]),
            "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "C", "Op", 20]]),
        })
        result = read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        assert "HR-Nomina" in result


class TestReaderResourceLimits:

    def test_rejects_excessive_rows(self, monkeypatch):
        import nexa_engine.modules.parametrizacion.shared.helpers.excel_reader as reader
        monkeypatch.setattr(reader, "MAX_EXCEL_ROWS_PER_SHEET", 3)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HR-Test"
        ws.append(["Col1"])
        for i in range(5):
            ws.append([f"val{i}"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        with pytest.raises(UploadError) as exc_info:
            read_excel_sheets(buf.getvalue(), "HR-")
        assert exc_info.value.code == "EXCEL_LIMIT_EXCEEDED"

    def test_rejects_cell_exceeding_max_length(self, monkeypatch):
        import nexa_engine.modules.parametrizacion.shared.helpers.excel_reader as reader
        monkeypatch.setattr(reader, "MAX_EXCEL_CELL_LENGTH", 10)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HR-Test"
        ws.append(["Col1"])
        ws.append(["A" * 20])  # 20 chars > 10
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        with pytest.raises(UploadError) as exc_info:
            read_excel_sheets(buf.getvalue(), "HR-")
        assert exc_info.value.code == "EXCEL_LIMIT_EXCEEDED"


# ===========================================================================
# 2. Value normalizer injection detection
# ===========================================================================

class TestCatalogByColumn:
    """Positive tests for CATALOG_BY_COLUMN sheet type (Req 4 + 8)."""

    def _read_hr_lv(self, rows) -> dict:
        from nexa_engine.modules.parametrizacion.hr.contracts import HR_CONTRACT
        from nexa_engine.modules.parametrizacion.shared.contracts.normalizer import normalize_sheets_by_contract
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                rows,
            ),
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cob", 1]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario"], [["Op", "Ag", 1]]),
            "HR-Recargos": (["Recargo", "Valor"], [["N", 0.3]]),
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["S", 0.08]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.08]]),
            "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "C", "Op", 20]]),
        })
        sheets = read_excel_sheets(content, "HR-", contract=HR_CONTRACT)
        normalized = normalize_sheets_by_contract(sheets, HR_CONTRACT)
        from nexa_engine.modules.parametrizacion.hr.mappers.mapper import HRMapper
        mapper = HRMapper()
        master = mapper.map("v1", normalized)
        return master.niveles.catalogs

    def test_catalog_columns_with_different_lengths_are_independent(self):
        """Each column forms its own catalog regardless of other columns' sizes.

        Tipo has 1 value, Rol has 3, Servicio has 2.
        Each must be extracted independently — no cross-row correlation.
        """
        rows = [
            ["Empleado",  "Agente",    "Cobranzas",  "Ces",  "Salud",  "RF"],
            [None,        "Supervisor","BPO",         None,   None,     None],
            [None,        "Director",  None,          None,   None,     None],
        ]
        catalogs = self._read_hr_lv(rows)

        # Independent lengths
        assert len(catalogs["tipo"]) == 1
        assert len(catalogs["rol"]) == 3
        assert len(catalogs["servicio"]) == 2

        # Correct values
        assert catalogs["tipo"] == [{"name": "Empleado"}]
        assert catalogs["rol"] == [
            {"name": "Agente"},
            {"name": "Supervisor"},
            {"name": "Director"},
        ]
        assert catalogs["servicio"] == [{"name": "Cobranzas"}, {"name": "BPO"}]

    def test_catalog_empty_cells_do_not_affect_other_columns(self):
        """Empty cells in one column must not displace or corrupt other columns."""
        rows = [
            ["Empleado",  "Agente",    "Cobranzas",  "Ces",  "Salud",  "RF"],
            [None,        None,         None,         None,  "Pensión", None],
            [None,        "Supervisor", None,         None,   None,     "RN"],
        ]
        catalogs = self._read_hr_lv(rows)

        assert catalogs["tipo"] == [{"name": "Empleado"}]
        assert catalogs["rol"] == [{"name": "Agente"}, {"name": "Supervisor"}]
        assert catalogs["servicio"] == [{"name": "Cobranzas"}]
        assert {"name": "Pensión"} in catalogs["ssparafiscales"]
        assert {"name": "RN"} in catalogs["recargo"]

    def test_catalog_values_output_as_name_objects_not_bare_strings(self):
        """Every catalog item is {\"name\": \"<value>\"}, not a bare string."""
        rows = [
            ["Empleado", "Agente", "Cobranzas", "Ces", "Salud", "RF"],
        ]
        catalogs = self._read_hr_lv(rows)

        for catalog_name, items in catalogs.items():
            for item in items:
                assert isinstance(item, dict), f"{catalog_name}: item must be dict, got {type(item)}"
                assert "name" in item, f"{catalog_name}: item must have 'name' key"
                assert isinstance(item["name"], str), f"{catalog_name}: name must be string"


class TestNormalizerInjection:

    def test_rejects_cell_starting_with_equals(self):
        sheets = {"HR-Test": [{"col1": "=SUM(A1:A10)"}]}
        with pytest.raises(ValidationError):
            normalize_all_sheets_values(sheets)

    def test_rejects_cell_starting_with_at(self):
        sheets = {"HR-Test": [{"col1": "@cmd /c evil"}]}
        with pytest.raises(ValidationError):
            normalize_all_sheets_values(sheets)

    def test_does_not_reject_negative_numbers(self):
        """Leading '-' in a numeric string is a valid negative number."""
        sheets = {"HR-Test": [{"col1": "-5.0"}]}
        result = normalize_all_sheets_values(sheets)
        assert result["HR-Test"][0]["col1"] == -5.0

    def test_does_not_reject_leading_plus_numeric(self):
        """Leading '+' followed by digits normalizes to a float."""
        sheets = {"HR-Test": [{"col1": "+3.14"}]}
        result = normalize_all_sheets_values(sheets)
        assert result["HR-Test"][0]["col1"] == pytest.approx(3.14)

    def test_does_not_coerce_si_to_bool(self):
        """'si' must NOT be coerced to True — it could be a catalog value."""
        sheets = {"HR-Test": [{"col1": "si"}]}
        result = normalize_all_sheets_values(sheets)
        # The normalizer now keeps plain strings as strings
        assert result["HR-Test"][0]["col1"] == "si"

    def test_does_not_coerce_no_to_bool(self):
        sheets = {"HR-Test": [{"col1": "no"}]}
        result = normalize_all_sheets_values(sheets)
        assert result["HR-Test"][0]["col1"] == "no"


# ===========================================================================
# 3. Full HTTP upload pipeline via TestClient
# ===========================================================================

def _install_gn_service(monkeypatch, tmp_path):
    import sys
    projects_root = Path(__file__).resolve().parents[4]
    if str(projects_root) not in sys.path:
        sys.path.insert(0, str(projects_root))

    from nexa_engine.db.providers.json_document_store import JsonDocumentStore
    from nexa_engine.modules.parametrizacion.gn.api import router as gn_router_module
    from nexa_engine.modules.parametrizacion.gn.mappers.gn_version_document_codec import GNVersionDocumentCodec
    from nexa_engine.modules.parametrizacion.gn.repositories.collections import GN_PARAMETRIZATION_COLLECTION
    from nexa_engine.modules.parametrizacion.gn.repositories.gn_repository import GNRepository
    from nexa_engine.modules.parametrizacion.gn.services.gn_service import GNService
    from nexa_engine.modules.parametrizacion.shared.repositories.version_index_repository import VersionIndexRepository

    store = JsonDocumentStore(tmp_path)
    repo = GNRepository(
        store=store,
        version_index_repository=VersionIndexRepository(store=store, collection=GN_PARAMETRIZATION_COLLECTION),
        codec=GNVersionDocumentCodec(),
    )
    repo.new_version_id = lambda: "gn-test"
    repo.now_iso = lambda: "2026-06-04T00:00:00Z"
    service = GNService(repository=repo)
    monkeypatch.setattr(gn_router_module, "_service", service)
    return gn_router_module


def _install_hr_service(monkeypatch, tmp_path):
    import sys
    projects_root = Path(__file__).resolve().parents[4]
    if str(projects_root) not in sys.path:
        sys.path.insert(0, str(projects_root))

    from nexa_engine.db.providers.json_document_store import JsonDocumentStore
    from nexa_engine.modules.parametrizacion.hr.api import router as hr_router_module
    from nexa_engine.modules.parametrizacion.hr.mappers.hr_version_document_codec import HRVersionDocumentCodec
    from nexa_engine.modules.parametrizacion.hr.repositories.collections import HR_PARAMETRIZATION_COLLECTION
    from nexa_engine.modules.parametrizacion.hr.repositories.hr_repository import HRRepository
    from nexa_engine.modules.parametrizacion.hr.services.hr_service import HRService
    from nexa_engine.modules.parametrizacion.shared.repositories.version_index_repository import VersionIndexRepository

    store = JsonDocumentStore(tmp_path)
    repo = HRRepository(
        store=store,
        version_index_repository=VersionIndexRepository(store=store, collection=HR_PARAMETRIZATION_COLLECTION),
        codec=HRVersionDocumentCodec(),
    )
    repo.new_version_id = lambda: "hr-test"
    repo.now_iso = lambda: "2026-06-04T00:00:00Z"
    service = HRService(repository=repo)
    monkeypatch.setattr(hr_router_module, "_service", service)
    return hr_router_module


def _install_op_service(monkeypatch, tmp_path):
    import sys
    projects_root = Path(__file__).resolve().parents[4]
    if str(projects_root) not in sys.path:
        sys.path.insert(0, str(projects_root))

    from nexa_engine.db.providers.json_document_store import JsonDocumentStore
    from nexa_engine.modules.parametrizacion.op.api import router as op_router_module
    from nexa_engine.modules.parametrizacion.op.mappers.op_version_document_codec import OPVersionDocumentCodec
    from nexa_engine.modules.parametrizacion.op.repositories.collections import OP_PARAMETRIZATION_COLLECTION
    from nexa_engine.modules.parametrizacion.op.repositories.op_repository import OPRepository
    from nexa_engine.modules.parametrizacion.op.services.op_service import OPService
    from nexa_engine.modules.parametrizacion.shared.repositories.version_index_repository import VersionIndexRepository

    store = JsonDocumentStore(tmp_path)
    repo = OPRepository(
        store=store,
        version_index_repository=VersionIndexRepository(store=store, collection=OP_PARAMETRIZATION_COLLECTION),
        codec=OPVersionDocumentCodec(),
    )
    repo.new_version_id = lambda: "op-test"
    repo.now_iso = lambda: "2026-06-04T00:00:00Z"
    service = OPService(repository=repo)
    monkeypatch.setattr(op_router_module, "_service", service)
    return op_router_module


@pytest.fixture
def isolated_app() -> FastAPI:
    return FastAPI()


class TestFullPipelinePositive:
    """Happy-path tests — valid files are accepted and persisted correctly."""

    def test_accepts_valid_gn_workbook(self, monkeypatch, tmp_path, isolated_app):
        mod = _install_gn_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/gn/upload",
            files={"file": ("GN.xlsx", _valid_gn_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_accepts_valid_hr_workbook(self, monkeypatch, tmp_path, isolated_app):
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", _valid_hr_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_accepts_valid_op_workbook(self, monkeypatch, tmp_path, isolated_app):
        mod = _install_op_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/op/upload",
            files={"file": ("OP.xlsx", _valid_op_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_hr_catalog_columns_are_independent(self, monkeypatch, tmp_path, isolated_app):
        """Each HR-LV column produces an independent catalog list."""
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        # Tipo has 1 value, Rol has 2, Servicio has 1, rest have 1 each
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [
                    ["Empleado", "Agente",      "Cobranzas", "Cesantías", "Salud", "RF"],
                    [None,       "Supervisor",  None,        None,        None,    None],
                ],
            ),
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cobranzas", 1]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario"], [["Op", "Ag", 1]]),
            "HR-Recargos": (["Recargo", "Valor"], [["N", 0.3]]),
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["S", 0.08]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.08]]),
            "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "C", "Op", 20]]),
        })
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["success"] is True

        # Verify persisted catalogs
        import json
        version_file = tmp_path / "hr" / "hr-test.json"
        payload = json.loads(version_file.read_text())
        catalogs = payload["niveles"]["catalogs"]

        assert catalogs["tipo"] == [{"name": "Empleado"}]
        assert catalogs["rol"] == [{"name": "Agente"}, {"name": "Supervisor"}]
        assert catalogs["servicio"] == [{"name": "Cobranzas"}]

    def test_hr_catalog_items_have_name_key(self, monkeypatch, tmp_path, isolated_app):
        """Every catalog item must be {\"name\": value}, never a bare string."""
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", _valid_hr_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        import json
        payload = json.loads((tmp_path / "hr" / "hr-test.json").read_text())
        for cat_list in payload["niveles"]["catalogs"].values():
            for item in cat_list:
                assert isinstance(item, dict), "Catalog item must be a dict"
                assert "name" in item, "Catalog item must have 'name' key"

    def test_invalid_file_does_not_create_version(self, monkeypatch, tmp_path, isolated_app):
        """A file that fails contract validation must not persist any version."""
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)

        bad_content = _workbook_bytes({
            "HR-LV": (["Tipo", "Rol"], [["E", "A"]]),  # missing required headers + sheets
        })
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR_bad.xlsx", bad_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert resp.json()["success"] is False

        # Storage directory must not have been written
        versions_file = tmp_path / "hr" / "versions.json"
        assert not versions_file.exists(), "versions.json must not exist after failed upload"

    def test_no_unauthorized_catalogs_wrapper(self, monkeypatch, tmp_path, isolated_app):
        """GN output must use 'catalogs' wrapper only as already defined in the contract."""
        mod = _install_gn_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/gn/upload",
            files={"file": ("GN.xlsx", _valid_gn_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        import json
        payload = json.loads((tmp_path / "gn" / "gn-test.json").read_text())
        # GN-LV is a catalog sheet and should have catalogs under lv.catalogs
        assert "lv" in payload
        assert "catalogs" in payload["lv"]


class TestFullPipelineNegative:
    """File-level rejection via full HTTP pipeline."""

    def test_rejects_xls_extension(self, monkeypatch, tmp_path, isolated_app):
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR_test.xls", b"not excel", "application/vnd.ms-excel")},
        )
        assert resp.status_code == 400

    def test_rejects_non_excel_bytes_with_xlsx_name(self, monkeypatch, tmp_path, isolated_app):
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR_test.xlsx", b"This is not a ZIP file at all", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is False
        assert body["error"]["code"] in ("INVALID_EXCEL_FILE", "UPLOAD_ERROR")

    def test_rejects_hr_missing_required_sheet_via_http(self, monkeypatch, tmp_path, isolated_app):
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        content = _workbook_bytes({
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cob", 1]]),
        })
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is False
        assert body["error"]["code"] == "VALIDATION_ERROR"
        # Details must mention the missing sheet
        details = " ".join(str(d) for d in (body["error"].get("details") or []))
        assert "HR-LV" in details

    def test_rejects_hr_wrong_header_via_http(self, monkeypatch, tmp_path, isolated_app):
        """'SS Parafiscales' (space, no &) is rejected — exact match required."""
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS Parafiscales", "Recargo"],  # wrong
                [["E", "A", "C", "Ces", "Sal", "RF"]],
            ),
        })
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is False
        assert body["error"]["code"] == "VALIDATION_ERROR"

    def test_invalid_numeric_value_blocks_persistence(self, monkeypatch, tmp_path, isolated_app):
        """A non-numeric value in a declared-numeric column raises ValidationError.

        HR-Recargos.Valor is typed PERCENTAGE_DECIMAL.  Sending the string
        'not_a_number' must cause a ValidationError that stops persistence.
        """
        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)

        content = _workbook_bytes({
            "HR-LV": (
                ["Tipo", "Rol", "Servicio", "Prestaciones", "SS&Parafiscales", "Recargo"],
                [["Empleado", "Agente", "Cobranzas", "Ces", "Salud", "RF"]],
            ),
            "HR-SalarioBasico": (["Servicio", "Valor"], [["Cobranzas", 1423000]]),
            "HR-Nomina": (["Tipo", "Rol", "Salario"], [["Op", "Ag", 1423000]]),
            "HR-Recargos": (["Recargo", "Valor"], [["Nocturno", "not_a_number"]]),  # ← invalid
            "HR-SegSocial": (["SS&Parafiscales", "Proporcion"], [["Salud", 0.085]]),
            "HR-Prestaciones": (["Prestaciones", "Valor"], [["Ces", 0.0833]]),
            "HR-Ratios": (["Cargo", "CategoriaServicio", "Tipo", "Agentes"], [["Sup", "Cob", "Op", 20]]),
        })
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is False
        # Error code can be VALIDATION_ERROR (contract normalizer) or DOMAIN_ERROR
        assert body["error"]["code"] in ("VALIDATION_ERROR", "DOMAIN_ERROR")

        # Storage must be empty — no partial write
        assert not (tmp_path / "hr" / "versions.json").exists(), (
            "versions.json must not exist when upload fails due to invalid numeric value"
        )
        assert not (tmp_path / "hr" / "hr-test.json").exists(), (
            "version payload must not be written when upload fails"
        )

    def test_rejects_file_too_large_via_http(self, monkeypatch, tmp_path, isolated_app):
        import nexa_engine.modules.parametrizacion.hr.api.router as hr_router_module_inst
        import nexa_engine.modules.shared.config.config as cfg
        monkeypatch.setattr(cfg, "MAX_EXCEL_UPLOAD_BYTES", 10)
        # Reload the router's imported constant
        import importlib
        import nexa_engine.modules.parametrizacion.hr.api.router as hr_router_reimport
        monkeypatch.setattr(hr_router_reimport, "MAX_EXCEL_UPLOAD_BYTES", 10)

        mod = _install_hr_service(monkeypatch, tmp_path)
        isolated_app.include_router(mod.router)
        client = TestClient(isolated_app)
        resp = client.post(
            "/parametrization/hr/upload",
            files={"file": ("HR.xlsx", b"A" * 20, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        body = resp.json()
        assert body["success"] is False
        assert body["error"]["code"] == "EXCEL_LIMIT_EXCEEDED"
