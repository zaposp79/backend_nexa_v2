"""Tests for OP-ICA rate guardrails.

Validates all 11 acceptance criteria from the ICA guardrail task.

ICA guardrail rules:
  * ICA == "Tasa" and Valor > MAX_TASA_ICA_DECIMAL (0.05) → ERROR (blocks upload)
  * ICA != "Tasa" and Valor > MAX_TASA_ICA_DECIMAL → WARNING only (non-blocking)
  * MAX_TASA_ICA_DECIMAL is configurable via environment variable
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pytest

from nexa_engine.modules.parametrizacion.op.validators.validator import OPValidator
from nexa_engine.modules.shared.exceptions import ValidationError

BACKEND_ROOT = Path(__file__).resolve().parents[3]


# ---------------------------------------------------------------------------
# OPValidator unit tests
# ---------------------------------------------------------------------------

class TestICAGuardrailTasaRows:
    """ICA == 'Tasa' rows: hard error when valor > threshold."""

    def _validate(self, rows):
        sheets = {"OP-ICA": rows}
        return OPValidator().validate(sheets)

    # 1. Armenia 0.6 must fail
    def test_tasa_0_6_armenia_is_error(self):
        result = self._validate([{"ciudad": "Armenia", "ica": "Tasa", "valor": 0.6}])
        assert not result.is_valid
        assert any("Armenia" in e for e in result.errors)
        assert any("INVALID_ICA_RATE" in e for e in result.errors)

    # 2. Valid Barranquilla 0.0125 must pass
    def test_tasa_0_0125_barranquilla_passes(self):
        result = self._validate([{"ciudad": "Barranquilla", "ica": "Tasa", "valor": 0.0125}])
        assert result.is_valid
        assert result.errors == []

    # 3. Exactly at threshold 0.05 must pass (not strictly greater)
    def test_tasa_exactly_at_threshold_passes(self):
        result = self._validate([{"ciudad": "TestCity", "ica": "Tasa", "valor": 0.05}])
        assert result.is_valid
        assert result.errors == []

    # 4. Marginally above threshold 0.0501 must fail
    def test_tasa_0_0501_just_above_threshold_fails(self):
        result = self._validate([{"ciudad": "TestCity", "ica": "Tasa", "valor": 0.0501}])
        assert not result.is_valid
        assert any("INVALID_ICA_RATE" in e for e in result.errors)

    # Boundary: exactly 0.05 must pass; 0.0501 must fail
    def test_threshold_boundary(self):
        at_threshold = self._validate([{"ciudad": "X", "ica": "Tasa", "valor": 0.05}])
        above_threshold = self._validate([{"ciudad": "X", "ica": "Tasa", "valor": 0.0501}])
        assert at_threshold.is_valid
        assert not above_threshold.is_valid

    def test_tasa_bogota_0_0097_passes(self):
        result = self._validate([{"ciudad": "Bogotá", "ica": "Tasa", "valor": 0.0097}])
        assert result.is_valid

    def test_tasa_manizales_0_0045_passes(self):
        result = self._validate([{"ciudad": "Manizales", "ica": "Tasa", "valor": 0.0045}])
        assert result.is_valid

    def test_tasa_case_insensitive_lower(self):
        """'tasa' (lowercase) also triggers the hard-error rule."""
        result = self._validate([{"ciudad": "Armenia", "ica": "tasa", "valor": 0.6}])
        assert not result.is_valid

    def test_tasa_case_insensitive_mixed(self):
        """'Tasa' (mixed case) also triggers the hard-error rule."""
        result = self._validate([{"ciudad": "Armenia", "ica": "Tasa", "valor": 0.6}])
        assert not result.is_valid

    def test_multiple_cities_one_anomalous(self):
        """Multiple cities: one anomalous Tasa row blocks the whole upload."""
        result = self._validate([
            {"ciudad": "Bogotá",      "ica": "Tasa", "valor": 0.0097},   # valid
            {"ciudad": "Barranquilla","ica": "Tasa", "valor": 0.0125},   # valid
            {"ciudad": "Armenia",     "ica": "Tasa", "valor": 0.6},      # INVALID
            {"ciudad": "Manizales",   "ica": "Tasa", "valor": 0.0045},   # valid
        ])
        assert not result.is_valid
        errors = " ".join(result.errors)
        assert "Armenia" in errors
        assert "INVALID_ICA_RATE" in errors

    def test_multiple_tasa_anomalies_all_reported(self):
        """Each anomalous Tasa row produces its own error entry."""
        result = self._validate([
            {"ciudad": "Armenia",  "ica": "Tasa", "valor": 0.6},
            {"ciudad": "FakeCity", "ica": "Tasa", "valor": 0.99},
        ])
        assert not result.is_valid
        assert len(result.errors) >= 2


class TestICAGuardrailSubcategoryRows:
    """ICA != 'Tasa' rows: warning only (non-blocking)."""

    def _validate(self, rows):
        sheets = {"OP-ICA": rows}
        return OPValidator().validate(sheets)

    # 5. Bomberos 0.1 must NOT fail — warning only
    def test_bomberos_0_1_bucaramanga_warning_not_error(self):
        result = self._validate([{"ciudad": "Bucaramanga", "ica": "Bomberos", "valor": 0.1}])
        assert result.is_valid        # not an error — non-blocking
        assert result.warnings        # but a warning is emitted
        assert any("Bomberos" in w or "bucaramanga" in w.lower() for w in result.warnings)

    def test_avisos_tableros_0_15_pereira_warning_not_error(self):
        result = self._validate([
            {"ciudad": "Pereira", "ica": "Avisos & Tableros", "valor": 0.15}
        ])
        assert result.is_valid
        assert result.warnings

    def test_subcategory_below_threshold_no_warning(self):
        result = self._validate([
            {"ciudad": "Bogotá", "ica": "Bomberos", "valor": 0.005}
        ])
        assert result.is_valid
        assert result.warnings == []

    def test_all_production_op_ica_subcategory_outliers_are_warnings(self):
        """All production subcategory values > threshold generate warnings, not errors."""
        result = self._validate([
            {"ciudad": "Bucaramanga",  "ica": "Bomberos",         "valor": 0.1},
            {"ciudad": "Cartagena",    "ica": "Bomberos",         "valor": 0.07},
            {"ciudad": "Pereira",      "ica": "Avisos & Tableros","valor": 0.15},
            {"ciudad": "Pereira",      "ica": "Bomberos",         "valor": 0.06},
            {"ciudad": "Santa Marta",  "ica": "Bomberos",         "valor": 0.07},
        ])
        assert result.is_valid               # non-blocking
        assert len(result.warnings) == 5     # one per anomalous row
        assert result.errors == []           # no errors

    def test_mixed_tasa_valid_and_subcategory_anomalous(self):
        """Valid Tasa row + anomalous Bomberos row: upload succeeds with warning."""
        result = self._validate([
            {"ciudad": "Bogotá",      "ica": "Tasa",     "valor": 0.0097},  # valid tasa
            {"ciudad": "Bogotá",      "ica": "Bomberos", "valor": 0.1},     # high but subcategory
        ])
        assert result.is_valid
        assert len(result.warnings) == 1
        assert result.errors == []


class TestICAGuardrailConfigurable:
    """MAX_TASA_ICA_DECIMAL can be overridden via monkeypatching config."""

    def test_custom_threshold_via_monkeypatch(self, monkeypatch):
        import nexa_engine.modules.parametrizacion.op.validators.validator as v_mod
        monkeypatch.setattr(v_mod, "MAX_TASA_ICA_DECIMAL", 0.02)

        sheets = {"OP-ICA": [
            {"ciudad": "Bogotá",     "ica": "Tasa", "valor": 0.0097},  # < 0.02 → ok
            {"ciudad": "Barranquilla","ica": "Tasa", "valor": 0.0125}, # < 0.02 → ok
        ]}
        result = OPValidator().validate(sheets)
        assert result.is_valid

    def test_strict_threshold_blocks_barranquilla(self, monkeypatch):
        import nexa_engine.modules.parametrizacion.op.validators.validator as v_mod
        monkeypatch.setattr(v_mod, "MAX_TASA_ICA_DECIMAL", 0.01)

        sheets = {"OP-ICA": [
            {"ciudad": "Barranquilla", "ica": "Tasa", "valor": 0.0125},  # > 0.01
        ]}
        result = OPValidator().validate(sheets)
        assert not result.is_valid
        assert any("Barranquilla" in e for e in result.errors)


# ---------------------------------------------------------------------------
# Full upload pipeline tests (via OPService)
# ---------------------------------------------------------------------------

def _make_op_workbook(ica_rows, poliza_rows=None) -> bytes:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    ws_ica = wb.create_sheet("OP-ICA")
    ws_ica.append(["Ciudad", "ICA", "Valor"])
    for row in ica_rows:
        ws_ica.append(row)

    ws_poliza = wb.create_sheet("OP-Poliza")
    ws_poliza.append(["Poliza", "Porcentaje", "PorcentajeExigido"])
    for row in (poliza_rows or [["Cumplimiento", 0.0062, 0.005]]):
        ws_poliza.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _op_service(tmp_path):
    from nexa_engine.db.providers.json_document_store import JsonDocumentStore
    from nexa_engine.modules.parametrizacion.op.mappers.op_version_document_codec import OPVersionDocumentCodec
    from nexa_engine.modules.parametrizacion.op.repositories.collections import OP_PARAMETRIZATION_COLLECTION
    from nexa_engine.modules.parametrizacion.op.repositories.op_repository import OPRepository
    from nexa_engine.modules.parametrizacion.op.services.op_service import OPService
    from nexa_engine.modules.parametrizacion.shared.repositories.version_index_repository import VersionIndexRepository

    store = JsonDocumentStore(tmp_path)
    repo = OPRepository(
        store=store,
        version_index_repository=VersionIndexRepository(store=store, collection=OP_PARAMETRIZATION_COLLECTION),
        codec=OPVersionDocumentCodec(),
    )
    repo.new_version_id = lambda: "op-test"
    repo.now_iso = lambda: "2026-06-04T00:00:00Z"
    return OPService(repository=repo)


class TestICAGuardrailPipeline:
    """End-to-end pipeline tests through OPService."""

    def test_armenia_tasa_0_6_blocks_upload(self, tmp_path):
        """Armenia Tasa=0.6 must block the upload — ValidationError raised."""
        service = _op_service(tmp_path)
        content = _make_op_workbook([
            ["Armenia",     "Tasa", 0.6],    # ANOMALY
            ["Bogotá",      "Tasa", 0.0097],
        ])
        with pytest.raises(ValidationError) as exc_info:
            service.process_upload("OP.xlsx", content)

        assert any("INVALID_ICA_RATE" in e for e in exc_info.value.errors)
        # Nothing persisted
        assert not (tmp_path / "op" / "op-test.json").exists()

    def test_valid_tasa_upload_succeeds(self, tmp_path):
        """Valid tasa values allow the upload to succeed."""
        service = _op_service(tmp_path)
        content = _make_op_workbook([
            ["Bogotá",       "Tasa", 0.0097],
            ["Barranquilla", "Tasa", 0.0125],
            ["Manizales",    "Tasa", 0.0045],
        ])
        resp = service.process_upload("OP.xlsx", content)
        assert resp.version_id == "op-test"
        assert (tmp_path / "op" / "op-test.json").exists()

    def test_armenia_corrected_to_0_006_passes(self, tmp_path):
        """Armenia corrected to 0.006 (0.6%) passes the guardrail."""
        service = _op_service(tmp_path)
        content = _make_op_workbook([
            ["Armenia",      "Tasa", 0.006],   # corrected value
            ["Bogotá",       "Tasa", 0.0097],
            ["Barranquilla", "Tasa", 0.0125],
        ])
        resp = service.process_upload("OP.xlsx", content)
        assert resp.version_id == "op-test"
        assert (tmp_path / "op" / "op-test.json").exists()
        # No ICA errors in warnings either
        assert not any("INVALID_ICA_RATE" in w for w in resp.warnings)

    def test_subcategory_above_threshold_warning_but_upload_succeeds(self, tmp_path):
        """Bomberos > threshold: warning but upload succeeds."""
        service = _op_service(tmp_path)
        content = _make_op_workbook([
            ["Bogotá",      "Tasa",     0.0097],  # valid
            ["Bucaramanga", "Bomberos", 0.1],     # warning
        ])
        resp = service.process_upload("OP.xlsx", content)
        assert resp.version_id == "op-test"  # upload succeeded
        assert any("Bomberos" in w or "Bucaramanga" in w for w in resp.warnings)

    def test_tasa_exactly_at_threshold_passes_pipeline(self, tmp_path):
        """Tasa == 0.05 (exactly at threshold) succeeds."""
        service = _op_service(tmp_path)
        content = _make_op_workbook([["TestCity", "Tasa", 0.05]])
        resp = service.process_upload("OP.xlsx", content)
        assert resp.version_id == "op-test"


# ---------------------------------------------------------------------------
# Production file tests
# ---------------------------------------------------------------------------

class TestProductionFileResults:
    """Test the actual production OP Excel file against the guardrail."""

    PRODUCTION_OP = BACKEND_ROOT / "excel/OP_productiva_2026-05-11-10-35-25.xlsx"

    def test_production_op_file_upload_succeeds(self, tmp_path):
        """The production OP Excel now succeeds — Armenia fixed from 0.6 to 0.006.

        Historical note: Previously Armenia Tasa=0.6 (60%) blocked upload with
        INVALID_ICA_RATE. This was corrected to 0.006 (0.6%), consistent with
        other municipal ICA rates (Bogotá 0.0097, Barranquilla 0.0125, etc.).
        """
        if not self.PRODUCTION_OP.exists():
            pytest.skip("Production OP Excel not available")

        service = _op_service(tmp_path)
        file_bytes = self.PRODUCTION_OP.read_bytes()

        # After Armenia fix (0.6 → 0.006), upload succeeds
        resp = service.process_upload("OP_productiva.xlsx", file_bytes)

        assert resp.version_id == "op-test"
        assert (tmp_path / "op" / "op-test.json").exists()
        # Armenia now passes — no ICA errors
        assert not any("Armenia" in w and "INVALID_ICA_RATE" in w for w in resp.warnings)

    def test_programmatic_armenia_fix_validation(self, tmp_path):
        """Verify that programmatically fixing Armenia 0.6 → 0.006 works (regression test).

        This documents the fix we applied to the production OP Excel.
        The production file OP_productiva_2026-05-11-10-35-25.xlsx now has
        Armenia corrected from 0.6 (60%) to 0.006 (0.6%).
        """
        if not self.PRODUCTION_OP.exists():
            pytest.skip("Production OP Excel not available")

        # Read original sheets and simulate Armenia fix
        import openpyxl as xl
        wb_orig = xl.load_workbook(self.PRODUCTION_OP, read_only=True, data_only=True)
        sheets_data = {}
        for sheet_name in wb_orig.sheetnames:
            ws = wb_orig[sheet_name]
            sheets_data[sheet_name] = list(ws.iter_rows(values_only=True))
        wb_orig.close()

        # Build test workbook with Armenia fix applied
        wb_test = xl.Workbook()
        default = wb_test.active
        wb_test.remove(default)
        for sheet_name, rows in sheets_data.items():
            ws = wb_test.create_sheet(sheet_name)
            for row in rows:
                if sheet_name == "OP-ICA":
                    row = list(row)
                    if (len(row) >= 3 and
                            str(row[0]).strip().lower() == "armenia" and
                            str(row[1]).strip().lower() == "tasa"):
                        # If Armenia is still 0.6 (old value), verify our fix works
                        if row[2] == 0.6:
                            row[2] = 0.006
                ws.append(list(row) if not isinstance(row, list) else row)

        buf = io.BytesIO()
        wb_test.save(buf)
        wb_test.close()

        service = _op_service(tmp_path)
        resp = service.process_upload("OP_test.xlsx", buf.getvalue())
        assert resp.version_id == "op-test"
        assert (tmp_path / "op" / "op-test.json").exists()
        # No ICA errors after fix
        ica_errors = [e for e in resp.warnings if "INVALID_ICA_RATE" in e]
        assert ica_errors == [], f"Expected no ICA errors, got: {ica_errors}"
