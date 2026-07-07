"""Excel parsing utilities — supports both .xlsx (openpyxl) and .xls (xlrd).

Public API
----------
read_excel_sheets(file_bytes, sheet_prefix, contract=None)
    Main entry point.  Returns ``{sheet_name: [row_dicts]}`` with normalised
    column keys.  When *contract* is supplied the function additionally:

    * Validates that every sheet in the workbook whose name starts with
      *sheet_prefix* is listed in the contract's authorised sheet set.
    * Validates that every required sheet is present.
    * Validates raw column headers (stripped of whitespace, but otherwise
      exact — same case, accents, special chars) against the contract's
      defined headers.

    These checks run *before* any normalisation so that, for example,
    ``"Prestacion"`` (missing tilde) is rejected even though it would
    normalise to the same key as the correct ``"Prestación"``.

list_sheets(file_bytes)
    Returns all sheet names.  Used by the preflight layer.

read_all_sheets(file_bytes)
    Reads every sheet without prefix filter.  Used in testing and diagnostics.

Format detection
----------------
Both functions detect the file format by magic bytes:
* ``PK\\x03\\x04`` (ZIP) → .xlsx path via openpyxl.
* ``\\xD0\\xCF\\x11\\xE0`` (CFBF/OLE) → .xls path via xlrd.
"""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Dict, List, Optional

import openpyxl
import xlrd

from nexa_engine.modules.parametrizacion.shared.contracts.base import ModuleContract
from nexa_engine.modules.shared.exceptions import UploadError, ValidationError
from nexa_engine.modules.shared.config.config import (
    MAX_EXCEL_CELL_LENGTH,
    MAX_EXCEL_CELLS,
    MAX_EXCEL_COLUMNS_PER_SHEET,
    MAX_EXCEL_ROWS_PER_SHEET,
    MAX_EXCEL_SHEETS,
)

# Magic bytes used to select the parser
_XLSX_MAGIC = b"PK\x03\x04"
_CFBF_MAGIC = b"\xD0\xCF\x11\xE0"


def _is_xls(file_bytes: bytes) -> bool:
    return len(file_bytes) >= 4 and file_bytes[:4] == _CFBF_MAGIC


_ACCENT_RE = re.compile(r"[^\x00-\x7F]")


def _normalize_column(name: str) -> str:
    """Normalize a column header: lowercase, strip accents, spaces→underscore.

    Note: camelCase NOT split (preserves backwards compat with existing JSON
    storage).  E.g. ``"MesDesde"`` → ``"mesdesde"`` (not ``"mes_desde"``).
    """
    if not isinstance(name, str):
        name = str(name)
    nfkd = unicodedata.normalize("NFKD", name)
    ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
    ascii_str = ascii_str.lower().strip()
    ascii_str = re.sub(r"[\s\-/]+", "_", ascii_str)
    ascii_str = re.sub(r"[^\w]", "", ascii_str)
    return ascii_str


def _strip_header(h: object) -> Optional[str]:
    """Return header value stripped of whitespace, or None if empty/None."""
    if h is None:
        return None
    s = str(h).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Contract-based validation helpers
# ---------------------------------------------------------------------------

def _validate_contract(
    sheet_prefix: str,
    module_contract: ModuleContract,
    all_sheet_names: List[str],
    prefixed_sheet_names: List[str],
    raw_headers_per_sheet: Dict[str, List[Optional[str]]],
) -> None:
    """Validate sheet names and raw headers against *module_contract*.

    Raises :class:`ValidationError` with a list of structured error strings.
    All errors are collected before raising so the client gets a complete
    picture in a single response.
    """
    errors: List[str] = []

    # Sheets present in the file that carry the module prefix
    present_prefixed = set(prefixed_sheet_names)

    # 1. Required sheets must be present
    for required in module_contract.required_names:
        if required not in present_prefixed:
            errors.append(
                f"INVALID_EXCEL_SHEET | hoja requerida '{required}' no encontrada. "
                f"Hojas con prefijo '{sheet_prefix}' en el archivo: "
                f"{sorted(present_prefixed) or '(ninguna)'}."
            )

    # 2. No unauthorised extra sheets allowed
    authorized = module_contract.authorized_names
    for name in present_prefixed:
        if name not in authorized:
            errors.append(
                f"INVALID_EXCEL_SHEET | hoja no autorizada '{name}'. "
                f"Hojas autorizadas: {sorted(authorized)}."
            )

    # 3. Per-sheet header validation (only for authorized + present sheets)
    for sheet_name, raw_hdrs in raw_headers_per_sheet.items():
        sheet_contract = module_contract.get_sheet(sheet_name)
        if sheet_contract is None:
            continue  # already caught above as unauthorised

        expected = sheet_contract.headers
        allow_trailing_unnamed = sheet_contract.allow_trailing_unnamed

        # Split raw headers into named (non-None) and trailing unnamed sections
        named_raw = [h for h in raw_hdrs if h is not None]
        # If trailing unnamed is allowed, only keep named headers for comparison
        # but verify that named headers beyond the expected count don't exist
        if allow_trailing_unnamed:
            # Find the last non-None header position to detect named extras
            last_named_idx = max(
                (i for i, h in enumerate(raw_hdrs) if h is not None),
                default=-1,
            )
            named_raw = [raw_hdrs[i] for i in range(last_named_idx + 1) if raw_hdrs[i] is not None]

        # Check count
        if len(named_raw) != len(expected):
            errors.append(
                f"INVALID_EXCEL_HEADER | hoja '{sheet_name}': se esperaban "
                f"{len(expected)} columnas nombradas, se encontraron {len(named_raw)}. "
                f"Esperados: {expected}. "
                f"Recibidos: {named_raw}."
            )
            continue  # skip per-column check when counts differ

        # Check each header exactly (after whitespace strip)
        for pos, (received, expected_h) in enumerate(zip(named_raw, expected), start=1):
            if received != expected_h:
                errors.append(
                    f"INVALID_EXCEL_HEADER | hoja '{sheet_name}', columna {pos}: "
                    f"encabezado esperado '{expected_h}', recibido '{received}'."
                )

    if errors:
        raise ValidationError(
            f"El archivo no cumple el contrato del módulo '{module_contract.module}'.",
            errors=errors,
        )


# ---------------------------------------------------------------------------
# XLS (BIFF8/OLE) helpers
# ---------------------------------------------------------------------------

def _open_xls(file_bytes: bytes) -> xlrd.Book:
    """Open XLS workbook from bytes, suppressing xlrd verbose logging."""
    try:
        return xlrd.open_workbook(file_contents=file_bytes, logfile=io.StringIO())
    except Exception as exc:
        raise UploadError(f"No se pudo abrir el archivo .xls: {exc}") from exc


def _cell_value_xls(cell: xlrd.sheet.Cell, datemode: int) -> object:
    """Convert an xlrd Cell to a Python value compatible with openpyxl data_only output."""
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_TEXT:
        return cell.value
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        return int(v) if isinstance(v, float) and v.is_integer() else v
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, datemode)
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    return None  # XL_CELL_ERROR → treat as empty (same as openpyxl data_only=True)


def _read_xls_sheets(
    file_bytes: bytes,
    sheet_prefix: str,
    contract: Optional[ModuleContract],
) -> Dict[str, List[dict]]:
    """XLS (BIFF8/OLE) reader using xlrd.  Same output contract as the XLSX reader."""
    book = _open_xls(file_bytes)
    datemode = book.datemode
    all_sheet_names = book.sheet_names()

    if len(all_sheet_names) > MAX_EXCEL_SHEETS:
        raise UploadError(
            f"El archivo tiene {len(all_sheet_names)} hojas; "
            f"el máximo permitido es {MAX_EXCEL_SHEETS}.",
            code="EXCEL_LIMIT_EXCEEDED",
        )

    prefixed_names = [n for n in all_sheet_names if n.startswith(sheet_prefix)]
    raw_headers_per_sheet: Dict[str, List[Optional[str]]] = {}
    result: Dict[str, List[dict]] = {}
    total_cells = 0

    for sheet_name in all_sheet_names:
        if not sheet_name.startswith(sheet_prefix):
            continue

        ws = book.sheet_by_name(sheet_name)

        if ws.ncols > MAX_EXCEL_COLUMNS_PER_SHEET:
            raise UploadError(
                f"La hoja '{sheet_name}' tiene {ws.ncols} columnas; "
                f"el máximo es {MAX_EXCEL_COLUMNS_PER_SHEET}.",
                code="EXCEL_LIMIT_EXCEEDED",
            )

        if ws.nrows == 0:
            raw_headers_per_sheet[sheet_name] = []
            result[sheet_name] = []
            continue

        header_raw = [_cell_value_xls(ws.cell(0, c), datemode) for c in range(ws.ncols)]
        raw_stripped = [_strip_header(h) for h in header_raw]
        raw_headers_per_sheet[sheet_name] = raw_stripped

        headers_norm = [
            _normalize_column(h) if h is not None else f"col_{i}"
            for i, h in enumerate(raw_stripped)
        ]

        sheet_rows: List[dict] = []
        for r in range(1, ws.nrows):
            if r > MAX_EXCEL_ROWS_PER_SHEET:
                raise UploadError(
                    f"La hoja '{sheet_name}' supera el límite de "
                    f"{MAX_EXCEL_ROWS_PER_SHEET} filas.",
                    code="EXCEL_LIMIT_EXCEEDED",
                )

            row = [_cell_value_xls(ws.cell(r, c), datemode) for c in range(ws.ncols)]
            if all(v is None or str(v).strip() == "" for v in row):
                continue

            row_dict: dict = {}
            for header, value in zip(headers_norm, row):
                if isinstance(value, str) and len(value) > MAX_EXCEL_CELL_LENGTH:
                    raise UploadError(
                        f"La hoja '{sheet_name}' contiene una celda que supera "
                        f"los {MAX_EXCEL_CELL_LENGTH} caracteres permitidos.",
                        code="EXCEL_LIMIT_EXCEEDED",
                    )
                row_dict[header] = value
                total_cells += 1

            if total_cells > MAX_EXCEL_CELLS:
                raise UploadError(
                    f"El archivo supera el límite de {MAX_EXCEL_CELLS} celdas.",
                    code="EXCEL_LIMIT_EXCEEDED",
                )

            sheet_rows.append(row_dict)

        result[sheet_name] = sheet_rows

    if contract is not None:
        _validate_contract(
            sheet_prefix=sheet_prefix,
            module_contract=contract,
            all_sheet_names=all_sheet_names,
            prefixed_sheet_names=prefixed_names,
            raw_headers_per_sheet=raw_headers_per_sheet,
        )

    return result


# ---------------------------------------------------------------------------
# Public functions
# ---------------------------------------------------------------------------

def list_sheets(file_bytes: bytes) -> List[str]:
    """Return all sheet names present in the workbook (.xlsx or .xls).

    Raises :class:`UploadError` if the bytes cannot be parsed.
    """
    if _is_xls(file_bytes):
        return _open_xls(file_bytes).sheet_names()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as exc:
        raise UploadError(f"No se pudo abrir el archivo Excel: {exc}") from exc


def read_excel_sheets(
    file_bytes: bytes,
    sheet_prefix: str,
    contract: Optional[ModuleContract] = None,
) -> Dict[str, List[dict]]:
    """Read sheets starting with *sheet_prefix* and return normalised row dicts.

    Format is detected from magic bytes — both .xlsx (OOXML/ZIP) and .xls
    (BIFF8/OLE) are supported.  When *contract* is provided, strict structural
    validation runs before any data is returned.

    Resource limits (from environment / :mod:`config`) are always enforced:

    * :data:`MAX_EXCEL_SHEETS` — total sheets in workbook.
    * :data:`MAX_EXCEL_ROWS_PER_SHEET` — rows per sheet.
    * :data:`MAX_EXCEL_COLUMNS_PER_SHEET` — columns per sheet.
    * :data:`MAX_EXCEL_CELLS` — total cells read.
    * :data:`MAX_EXCEL_CELL_LENGTH` — length of any single string cell.

    Raises
    ------
    :class:`UploadError`
        Workbook cannot be opened or a resource limit is exceeded.
    :class:`ValidationError`
        Contract violation (wrong sheets or headers).
    """
    if _is_xls(file_bytes):
        return _read_xls_sheets(file_bytes, sheet_prefix, contract)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise UploadError(f"No se pudo abrir el archivo Excel: {exc}") from exc

    try:
        all_sheet_names = wb.sheetnames

        # Limit: total sheets in workbook
        if len(all_sheet_names) > MAX_EXCEL_SHEETS:
            raise UploadError(
                f"El archivo tiene {len(all_sheet_names)} hojas; "
                f"el máximo permitido es {MAX_EXCEL_SHEETS}.",
                code="EXCEL_LIMIT_EXCEEDED",
            )

        prefixed_names = [n for n in all_sheet_names if n.startswith(sheet_prefix)]

        # Collect raw headers for contract validation
        raw_headers_per_sheet: Dict[str, List[Optional[str]]] = {}
        result: Dict[str, List[dict]] = {}
        total_cells = 0

        for sheet_name in all_sheet_names:
            if not sheet_name.startswith(sheet_prefix):
                continue

            ws = wb[sheet_name]

            # Limit: columns
            max_col = ws.max_column or 0
            if max_col > MAX_EXCEL_COLUMNS_PER_SHEET:
                raise UploadError(
                    f"La hoja '{sheet_name}' tiene {max_col} columnas; "
                    f"el máximo es {MAX_EXCEL_COLUMNS_PER_SHEET}.",
                    code="EXCEL_LIMIT_EXCEEDED",
                )

            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                raw_headers_per_sheet[sheet_name] = []
                result[sheet_name] = []
                continue

            # Strip whitespace from each raw header for comparison
            raw_stripped = [_strip_header(h) for h in header_row]
            raw_headers_per_sheet[sheet_name] = raw_stripped

            # Build normalised header list (None → col_N placeholder)
            headers_norm = [
                _normalize_column(h) if h is not None else f"col_{i}"
                for i, h in enumerate(raw_stripped)
            ]

            sheet_rows: List[dict] = []
            row_count = 0

            for row in rows_iter:
                # Limit: rows
                row_count += 1
                if row_count > MAX_EXCEL_ROWS_PER_SHEET:
                    raise UploadError(
                        f"La hoja '{sheet_name}' supera el límite de "
                        f"{MAX_EXCEL_ROWS_PER_SHEET} filas.",
                        code="EXCEL_LIMIT_EXCEEDED",
                    )

                # Skip entirely empty rows
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                row_dict: dict = {}
                for header, value in zip(headers_norm, row):
                    # Limit: cell length
                    if isinstance(value, str) and len(value) > MAX_EXCEL_CELL_LENGTH:
                        raise UploadError(
                            f"La hoja '{sheet_name}' contiene una celda que supera "
                            f"los {MAX_EXCEL_CELL_LENGTH} caracteres permitidos.",
                            code="EXCEL_LIMIT_EXCEEDED",
                        )
                    row_dict[header] = value
                    total_cells += 1

                # Limit: total cells
                if total_cells > MAX_EXCEL_CELLS:
                    raise UploadError(
                        f"El archivo supera el límite de {MAX_EXCEL_CELLS} celdas.",
                        code="EXCEL_LIMIT_EXCEEDED",
                    )

                sheet_rows.append(row_dict)

            result[sheet_name] = sheet_rows

        # Contract validation (after all headers collected, before returning data)
        if contract is not None:
            _validate_contract(
                sheet_prefix=sheet_prefix,
                module_contract=contract,
                all_sheet_names=all_sheet_names,
                prefixed_sheet_names=prefixed_names,
                raw_headers_per_sheet=raw_headers_per_sheet,
            )

        return result

    finally:
        wb.close()


def read_all_sheets(file_bytes: bytes) -> Dict[str, List[dict]]:
    """Like :func:`read_excel_sheets` but reads every sheet (no prefix filter).

    Supports both .xlsx and .xls — format detected from magic bytes.
    """
    if _is_xls(file_bytes):
        # Empty prefix matches all sheet names (every string starts with "")
        return _read_xls_sheets(file_bytes, "", None)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise UploadError(f"No se pudo abrir el archivo Excel: {exc}") from exc

    try:
        result: Dict[str, List[dict]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result[sheet_name] = []
                continue
            raw_headers = rows[0]
            headers = [
                _normalize_column(h) if h is not None else f"col_{i}"
                for i, h in enumerate(raw_headers)
            ]
            sheet_rows = []
            for row in rows[1:]:
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                row_dict = dict(zip(headers, row))
                sheet_rows.append(row_dict)
            result[sheet_name] = sheet_rows
        return result
    finally:
        wb.close()
